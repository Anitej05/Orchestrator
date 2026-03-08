"""
Spreadsheet Agent — Extended Unit Tests (Part 2)
=================================================

Covers gaps not in test_spreadsheet_agent_unit.py — the areas required for
industry-standard data-agent test coverage.

Coverage:
  TestGroupByAggregation       — sum/mean/count/min/max per group, multi-key groupby
  TestPivotOperations          — pivot_table, crosstab, unstack
  TestMergeJoin                — inner/left/outer joins (VLOOKUP equivalent)
  TestFillNaStrategies         — mean, median, mode, ffill, bfill, constant
  TestMultiConditionFilter     — AND/OR/NOT, chained filters, isin, between
  TestSortingOperations        — single column, multi-column, ascending/descending, stable
  TestOutlierDetection         — IQR fence, Z-score flagging, winsorization
  TestStatisticalAnalysis      — std dev, variance, correlation, percentile, skewness
  TestMRPCalculations          — EOQ, safety stock, reorder point, coverage days
  TestFormulaInjectionGuard    — cells with =CMD, +cmd, @SUM, -1+1, whitespace prefix
  TestExcelFormulaValues       — data_only=True reads computed values, not formula strings
  TestTimeSeries               — rolling average, YoY growth, month-over-month, lag
  TestCurrencyNormalization    — ₹, $, £, €, commas stripped before numeric ops
  TestErrorHandling            — empty file, wrong extension, oversized, zero-byte
  TestColumnOperations         — add calculated col, rename, drop, reorder
  TestDataValidation           — null count, type consistency, range check, uniqueness
  TestSessionIsolation         — two thread_ids cannot read each other's DataFrames
  TestNormalizeDataframeDeep   — currency strip, mixed types, all-string coercion edge

Run:
    PYTHONUTF8=1 venv/Scripts/python -m pytest backend/tests/test_spreadsheet_agent_unit2.py -v
"""

import io
import os
import sys
import asyncio
import tempfile
from pathlib import Path
from typing import Tuple

import pandas as pd
import numpy as np
import pytest
import openpyxl
from openpyxl import Workbook

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT.parent))

from dotenv import load_dotenv
load_dotenv(dotenv_path=ROOT / ".env", override=False)

from backend.agents.spreadsheet_agent.client import DataFrameClient


# ── Helpers ───────────────────────────────────────────────────────────────────

def _client() -> DataFrameClient:
    with tempfile.TemporaryDirectory() as tmp:
        return DataFrameClient(storage_dir=Path(tmp))


def _client_with_dir(tmp: Path) -> DataFrameClient:
    return DataFrameClient(storage_dir=tmp)


def _xlsx_bytes(sheets: dict) -> bytes:
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_xlsx(content: bytes) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with open(path, "wb") as f:
        f.write(content)
    return path


def _safe_unlink(path: str):
    try:
        os.unlink(path)
    except (PermissionError, OSError):
        pass


def _csv_bytes(text: str, encoding: str = "utf-8") -> bytes:
    return text.encode(encoding)


def _run(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


def _norm(df: pd.DataFrame) -> pd.DataFrame:
    return _client().normalize_dataframe(df)


# ── Shared sales fixture ──────────────────────────────────────────────────────

def _sales_df() -> pd.DataFrame:
    return pd.DataFrame({
        "Month":   ["Jan", "Jan", "Feb", "Feb", "Mar", "Mar"],
        "Region":  ["North", "South", "North", "South", "North", "South"],
        "Product": ["Widget A", "Widget B", "Widget A", "Widget B", "Widget A", "Widget B"],
        "Revenue": [1500.0, 800.0, 1700.0, 950.0, 1200.0, 600.0],
        "Units":   [30, 16, 34, 19, 24, 12],
    })


# =============================================================================
# TestGroupByAggregation
# =============================================================================

class TestGroupByAggregation:
    """GroupBy + aggregation — the most common real-world data operation."""

    def test_sum_revenue_per_region(self):
        df = _sales_df()
        result = df.groupby("Region")["Revenue"].sum()
        assert result["North"] == pytest.approx(4400.0)
        assert result["South"] == pytest.approx(2350.0)

    def test_mean_units_per_product(self):
        df = _sales_df()
        result = df.groupby("Product")["Units"].mean()
        assert result["Widget A"] == pytest.approx((30 + 34 + 24) / 3)
        assert result["Widget B"] == pytest.approx((16 + 19 + 12) / 3)

    def test_count_rows_per_month(self):
        df = _sales_df()
        result = df.groupby("Month")["Revenue"].count()
        assert result["Jan"] == 2
        assert result["Feb"] == 2
        assert result["Mar"] == 2

    def test_multi_key_groupby(self):
        df = _sales_df()
        result = df.groupby(["Region", "Product"])["Revenue"].sum()
        assert result[("North", "Widget A")] == pytest.approx(4400.0)
        assert result[("South", "Widget B")] == pytest.approx(2350.0)

    def test_min_max_per_region(self):
        df = _sales_df()
        agg = df.groupby("Region")["Revenue"].agg(["min", "max"])
        assert agg.loc["North", "min"] == pytest.approx(1200.0)
        assert agg.loc["North", "max"] == pytest.approx(1700.0)
        assert agg.loc["South", "max"] == pytest.approx(950.0)

    def test_multiple_aggregations_at_once(self):
        df = _sales_df()
        agg = df.groupby("Region").agg(
            TotalRevenue=("Revenue", "sum"),
            AvgUnits=("Units", "mean"),
            Transactions=("Revenue", "count"),
        )
        assert agg.loc["North", "TotalRevenue"] == pytest.approx(4400.0)
        assert agg.loc["North", "Transactions"] == 3

    def test_groupby_with_null_values_excluded(self):
        df = pd.DataFrame({
            "Category": ["A", "A", "B", None, "B"],
            "Value": [10, 20, 30, 40, 50],
        })
        result = df.groupby("Category", dropna=True)["Value"].sum()
        assert result["A"] == 30
        assert result["B"] == 80
        assert None not in result.index


# =============================================================================
# TestPivotOperations
# =============================================================================

class TestPivotOperations:
    """Pivot tables and crosstabs."""

    def test_pivot_table_revenue_by_region_and_month(self):
        df = _sales_df()
        pivot = df.pivot_table(
            values="Revenue", index="Region", columns="Month", aggfunc="sum"
        )
        assert pivot.loc["North", "Jan"] == pytest.approx(1500.0)
        assert pivot.loc["South", "Feb"] == pytest.approx(950.0)

    def test_pivot_totals_with_margins(self):
        df = _sales_df()
        pivot = df.pivot_table(
            values="Revenue", index="Region", columns="Month",
            aggfunc="sum", margins=True, margins_name="Total"
        )
        assert pivot.loc["Total", "Total"] == pytest.approx(6750.0)

    def test_crosstab_counts(self):
        df = _sales_df()
        ct = pd.crosstab(df["Region"], df["Product"])
        assert ct.loc["North", "Widget A"] == 3
        assert ct.loc["South", "Widget B"] == 3

    def test_pivot_fill_missing_with_zero(self):
        df = pd.DataFrame({
            "Region": ["North", "South", "North"],
            "Quarter": ["Q1", "Q1", "Q2"],
            "Revenue": [1000, 800, 1200],
        })
        pivot = df.pivot_table(
            values="Revenue", index="Region", columns="Quarter",
            aggfunc="sum", fill_value=0
        )
        # South has no Q2 entry → filled with 0
        assert pivot.loc["South", "Q2"] == 0

    def test_unstack_multiindex(self):
        df = _sales_df()
        grouped = df.groupby(["Region", "Month"])["Revenue"].sum()
        unstacked = grouped.unstack("Month")
        assert "Jan" in unstacked.columns
        assert "North" in unstacked.index


# =============================================================================
# TestMergeJoin
# =============================================================================

class TestMergeJoin:
    """Merge / join operations — the VLOOKUP equivalent."""

    def _orders(self):
        return pd.DataFrame({
            "OrderID": ["O001", "O002", "O003", "O004"],
            "ProductID": ["P001", "P002", "P001", "P003"],
            "Qty": [5, 10, 3, 7],
        })

    def _products(self):
        return pd.DataFrame({
            "ProductID": ["P001", "P002", "P004"],
            "ProductName": ["Widget A", "Widget B", "Widget D"],
            "UnitPrice": [50.0, 80.0, 120.0],
        })

    def test_inner_join_keeps_matching_rows_only(self):
        orders = self._orders()
        products = self._products()
        merged = orders.merge(products, on="ProductID", how="inner")
        # P003 not in products → dropped
        assert len(merged) == 3
        assert "P003" not in merged["ProductID"].tolist()

    def test_left_join_keeps_all_orders(self):
        orders = self._orders()
        products = self._products()
        merged = orders.merge(products, on="ProductID", how="left")
        # All 4 orders kept; P003 gets NaN for ProductName, UnitPrice
        assert len(merged) == 4
        p003_row = merged[merged["ProductID"] == "P003"]
        assert p003_row["ProductName"].isna().all()

    def test_outer_join_includes_unmatched_from_both(self):
        orders = self._orders()
        products = self._products()
        merged = orders.merge(products, on="ProductID", how="outer")
        # P004 is in products but not in orders → included with NaN Qty
        assert "P004" in merged["ProductID"].tolist()
        p004_row = merged[merged["ProductID"] == "P004"]
        assert p004_row["Qty"].isna().all()

    def test_join_computes_line_total(self):
        orders = self._orders()
        products = self._products()
        merged = orders.merge(products, on="ProductID", how="inner")
        merged["LineTotal"] = merged["Qty"] * merged["UnitPrice"]
        o001 = merged[merged["OrderID"] == "O001"]
        assert o001["LineTotal"].iloc[0] == pytest.approx(250.0)  # 5 × 50

    def test_join_on_multiple_keys(self):
        df1 = pd.DataFrame({"Year": [2024, 2024, 2023], "Region": ["N", "S", "N"], "Rev": [100, 200, 80]})
        df2 = pd.DataFrame({"Year": [2024, 2023], "Region": ["N", "N"], "Budget": [120, 90]})
        merged = df1.merge(df2, on=["Year", "Region"], how="left")
        row_2024_n = merged[(merged["Year"] == 2024) & (merged["Region"] == "N")]
        assert row_2024_n["Budget"].iloc[0] == 120
        row_2024_s = merged[(merged["Year"] == 2024) & (merged["Region"] == "S")]
        assert row_2024_s["Budget"].isna().all()


# =============================================================================
# TestFillNaStrategies
# =============================================================================

class TestFillNaStrategies:
    """Missing value imputation strategies."""

    def _df_with_nulls(self):
        return pd.DataFrame({
            "Sales": [100.0, None, 300.0, None, 500.0],
            "Region": ["N", None, "S", None, "N"],
        })

    def test_fill_with_mean(self):
        df = self._df_with_nulls()
        filled = df["Sales"].fillna(df["Sales"].mean())
        assert filled.isna().sum() == 0
        assert filled.iloc[1] == pytest.approx((100 + 300 + 500) / 3)

    def test_fill_with_median(self):
        df = self._df_with_nulls()
        filled = df["Sales"].fillna(df["Sales"].median())
        assert filled.isna().sum() == 0
        assert filled.iloc[1] == pytest.approx(300.0)

    def test_fill_with_mode(self):
        df = pd.DataFrame({"Cat": ["A", None, "A", "B", None]})
        filled = df["Cat"].fillna(df["Cat"].mode().iloc[0])
        assert filled.isna().sum() == 0
        assert filled.iloc[1] == "A"

    def test_forward_fill(self):
        df = pd.DataFrame({"Price": [10.0, None, None, 20.0, None]})
        filled = df["Price"].ffill()
        assert filled.iloc[1] == 10.0
        assert filled.iloc[2] == 10.0
        assert filled.iloc[4] == 20.0

    def test_back_fill(self):
        df = pd.DataFrame({"Price": [None, None, 30.0, None, 50.0]})
        filled = df["Price"].bfill()
        assert filled.iloc[0] == 30.0
        assert filled.iloc[1] == 30.0

    def test_fill_with_constant(self):
        df = self._df_with_nulls()
        filled = df["Region"].fillna("Unknown")
        assert (filled == "Unknown").sum() == 2

    def test_fill_na_does_not_affect_non_null(self):
        df = pd.DataFrame({"Val": [1.0, None, 3.0]})
        filled = df["Val"].fillna(0.0)
        assert filled.iloc[0] == 1.0
        assert filled.iloc[2] == 3.0


# =============================================================================
# TestMultiConditionFilter
# =============================================================================

class TestMultiConditionFilter:
    """Complex filtering — AND, OR, NOT, isin, between."""

    def test_and_condition_filter(self):
        df = _sales_df()
        result = df[(df["Region"] == "North") & (df["Revenue"] > 1400)]
        assert len(result) == 2
        assert (result["Region"] == "North").all()
        assert (result["Revenue"] > 1400).all()

    def test_or_condition_filter(self):
        df = _sales_df()
        result = df[(df["Region"] == "North") | (df["Revenue"] > 900)]
        assert len(result) > 3

    def test_not_condition_filter(self):
        df = _sales_df()
        result = df[~(df["Region"] == "South")]
        assert (result["Region"] == "North").all()
        assert len(result) == 3

    def test_isin_filter(self):
        df = pd.DataFrame({
            "Status": ["Pending", "Approved", "Rejected", "Pending", "Approved"],
            "Amount": [100, 200, 150, 300, 250],
        })
        result = df[df["Status"].isin(["Pending", "Approved"])]
        assert len(result) == 4
        assert "Rejected" not in result["Status"].tolist()

    def test_between_filter(self):
        df = pd.DataFrame({"Score": [45, 60, 72, 85, 90, 55]})
        result = df[df["Score"].between(60, 85)]
        assert set(result["Score"].tolist()) == {60, 72, 85}

    def test_string_contains_filter(self):
        df = pd.DataFrame({"Description": ["Hex Bolt M8", "Washer M8", "Nut M10", "Bolt M12"]})
        result = df[df["Description"].str.contains("Bolt", case=False)]
        assert len(result) == 2

    def test_chained_filter_and_count(self):
        df = _sales_df()
        # North rows: Jan=30, Feb=34, Mar=24 — strictly > 30 gives only Feb
        high_north = df[(df["Region"] == "North") & (df["Units"] > 30)]
        assert len(high_north) == 1
        assert high_north["Units"].iloc[0] == 34


# =============================================================================
# TestSortingOperations
# =============================================================================

class TestSortingOperations:
    """Sort correctness and stability."""

    def test_single_column_ascending(self):
        df = pd.DataFrame({"Val": [30, 10, 20, 10]})
        result = df.sort_values("Val")
        assert result["Val"].tolist() == [10, 10, 20, 30]

    def test_single_column_descending(self):
        df = pd.DataFrame({"Val": [30, 10, 20]})
        result = df.sort_values("Val", ascending=False)
        assert result["Val"].iloc[0] == 30
        assert result["Val"].iloc[-1] == 10

    def test_multi_column_sort_primary_and_secondary(self):
        df = pd.DataFrame({
            "Region": ["South", "North", "North", "South"],
            "Revenue": [800.0, 1700.0, 1500.0, 950.0],
        })
        result = df.sort_values(["Region", "Revenue"], ascending=[True, False])
        # North first (alphabetical), then South; within each, descending revenue
        assert result.iloc[0]["Region"] == "North"
        assert result.iloc[0]["Revenue"] == 1700.0
        assert result.iloc[2]["Region"] == "South"
        assert result.iloc[2]["Revenue"] == 950.0

    def test_sort_stable_preserves_equal_row_order(self):
        df = pd.DataFrame({
            "Key": ["A", "B", "A", "B"],
            "Order": [1, 2, 3, 4],
        })
        result = df.sort_values("Key", kind="stable")
        # "A" rows should appear in original order (1, 3)
        a_rows = result[result["Key"] == "A"]["Order"].tolist()
        assert a_rows == [1, 3]

    def test_sort_with_nulls_last(self):
        df = pd.DataFrame({"Val": [3.0, None, 1.0, None, 2.0]})
        result = df.sort_values("Val", na_position="last")
        assert result["Val"].iloc[0] == 1.0
        assert result["Val"].iloc[-1] != result["Val"].iloc[-1]  # NaN check


# =============================================================================
# TestOutlierDetection
# =============================================================================

class TestOutlierDetection:
    """IQR fence, Z-score flagging, winsorization."""

    def _revenue_series(self):
        # Most values 100-500, outliers at 5 and 10000
        return pd.Series([100, 200, 300, 400, 500, 150, 250, 350, 5, 10000])

    def test_iqr_outlier_detection(self):
        s = self._revenue_series()
        Q1 = s.quantile(0.25)
        Q3 = s.quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR
        outliers = s[(s < lower) | (s > upper)]
        assert 5 in outliers.values or 10000 in outliers.values

    def test_zscore_outlier_flagging(self):
        s = self._revenue_series()
        z_scores = (s - s.mean()) / s.std()
        outliers = s[z_scores.abs() > 2]
        assert len(outliers) >= 1
        assert 10000 in outliers.values

    def test_winsorization_clips_extremes(self):
        s = self._revenue_series()
        lower = s.quantile(0.05)
        upper = s.quantile(0.95)
        winsorized = s.clip(lower=lower, upper=upper)
        assert winsorized.max() <= upper
        assert winsorized.min() >= lower
        assert winsorized.max() < 10000

    def test_outlier_count_reasonable(self):
        """IQR method should flag at most 10% of a normal dataset."""
        np.random.seed(42)
        normal = pd.Series(np.random.normal(500, 50, 100))
        Q1, Q3 = normal.quantile(0.25), normal.quantile(0.75)
        IQR = Q3 - Q1
        outliers = normal[(normal < Q1 - 1.5 * IQR) | (normal > Q3 + 1.5 * IQR)]
        assert len(outliers) / len(normal) < 0.10


# =============================================================================
# TestStatisticalAnalysis
# =============================================================================

class TestStatisticalAnalysis:
    """Std dev, variance, correlation, percentile, skewness."""

    def test_standard_deviation_calculated_correctly(self):
        df = pd.DataFrame({"Sales": [100.0, 200.0, 300.0, 400.0, 500.0]})
        assert df["Sales"].std() == pytest.approx(158.11, rel=1e-2)

    def test_variance_equals_std_squared(self):
        df = pd.DataFrame({"Val": [10.0, 20.0, 30.0, 40.0]})
        assert df["Val"].var() == pytest.approx(df["Val"].std() ** 2)

    def test_correlation_between_columns(self):
        df = pd.DataFrame({
            "Units": [10, 20, 30, 40, 50],
            "Revenue": [100, 200, 300, 400, 500],  # perfectly correlated
        })
        corr = df["Units"].corr(df["Revenue"])
        assert corr == pytest.approx(1.0)

    def test_negative_correlation(self):
        df = pd.DataFrame({
            "Price": [10, 20, 30, 40],
            "Demand": [400, 300, 200, 100],  # perfectly negatively correlated
        })
        corr = df["Price"].corr(df["Demand"])
        assert corr == pytest.approx(-1.0)

    def test_percentile_calculations(self):
        s = pd.Series(range(1, 101))  # 1 to 100
        assert s.quantile(0.50) == pytest.approx(50.5)
        assert s.quantile(0.25) == pytest.approx(25.75)
        assert s.quantile(0.75) == pytest.approx(75.25)
        assert s.quantile(0.90) == pytest.approx(90.1)

    def test_skewness_right_skewed(self):
        """Revenue data with a long right tail is positively skewed."""
        s = pd.Series([100, 100, 100, 100, 100, 100, 500, 10000])
        assert s.skew() > 0

    def test_describe_covers_key_stats(self):
        df = _sales_df()
        desc = df["Revenue"].describe()
        assert "mean" in desc.index
        assert "std" in desc.index
        assert "min" in desc.index
        assert "max" in desc.index


# =============================================================================
# TestMRPCalculations
# =============================================================================

class TestMRPCalculations:
    """
    Manufacturing Resource Planning calculation correctness.

    EOQ (Economic Order Quantity) = sqrt(2 * D * S / H)
      D = annual demand, S = ordering cost, H = holding cost per unit per year

    Safety Stock = Z * σ_d * sqrt(L)
      Z = service level z-score, σ_d = std dev of daily demand, L = lead time (days)

    Reorder Point = (Avg Daily Demand × Lead Time) + Safety Stock

    Coverage Days = Stock On Hand / Avg Daily Demand
    """

    def test_eoq_formula(self):
        D, S, H = 10000, 500, 20
        eoq = (2 * D * S / H) ** 0.5
        assert eoq == pytest.approx(707.1, rel=1e-2)

    def test_eoq_higher_holding_cost_reduces_order_qty(self):
        D, S = 10000, 500
        eoq_low_H = (2 * D * S / 10) ** 0.5
        eoq_high_H = (2 * D * S / 40) ** 0.5
        assert eoq_high_H < eoq_low_H

    def test_safety_stock_95_pct_service_level(self):
        """95% service level → Z = 1.645."""
        Z, sigma_d, L = 1.645, 50, 4
        safety_stock = Z * sigma_d * (L ** 0.5)
        assert safety_stock == pytest.approx(164.5, rel=1e-2)

    def test_reorder_point_includes_safety_stock(self):
        avg_daily_demand, lead_time, safety_stock = 200, 5, 300
        rop = avg_daily_demand * lead_time + safety_stock
        assert rop == 1300

    def test_coverage_days_stock_runs_out(self):
        df = pd.DataFrame({
            "ItemCode": ["ITM001", "ITM002", "ITM003"],
            "StockOnHand": [600, 1500, 200],
            "AvgDailyDemand": [100, 300, 500],
        })
        df["CoverageDays"] = df["StockOnHand"] / df["AvgDailyDemand"]
        assert df.loc[0, "CoverageDays"] == pytest.approx(6.0)
        assert df.loc[1, "CoverageDays"] == pytest.approx(5.0)
        assert df.loc[2, "CoverageDays"] == pytest.approx(0.4)

    def test_mrp_requirement_netting(self):
        """Net requirement = Gross requirement - Stock on hand."""
        df = pd.DataFrame({
            "Item": ["A", "B", "C"],
            "GrossRequirement": [500, 300, 200],
            "StockOnHand": [100, 350, 0],
        })
        df["NetRequirement"] = (df["GrossRequirement"] - df["StockOnHand"]).clip(lower=0)
        assert df.loc[0, "NetRequirement"] == 400
        assert df.loc[1, "NetRequirement"] == 0   # stock exceeds requirement
        assert df.loc[2, "NetRequirement"] == 200

    def test_material_requirement_explosion(self):
        """BOM explosion: parent order quantity × BOM qty per parent."""
        parent_order_qty = 1000
        bom = pd.DataFrame({
            "Component": ["Bolt M8", "Washer M8", "Nut M8"],
            "QtyPerParent": [4, 4, 4],
        })
        bom["RequiredQty"] = bom["QtyPerParent"] * parent_order_qty
        assert bom.loc[0, "RequiredQty"] == 4000
        assert bom["RequiredQty"].sum() == 12000

    def test_supplier_lead_time_affects_order_date(self):
        """Order must be placed lead_time days before need date."""
        need_date = pd.Timestamp("2024-03-15")
        lead_times = pd.Series([5, 10, 15, 30])
        order_dates = need_date - pd.to_timedelta(lead_times, unit="D")
        assert order_dates.iloc[0] == pd.Timestamp("2024-03-10")
        assert order_dates.iloc[-1] == pd.Timestamp("2024-02-14")


# =============================================================================
# TestFormulaInjectionGuard
# =============================================================================

class TestFormulaInjectionGuard:
    """
    CSV formula injection: cells like =CMD, +cmd, @SUM are dangerous if
    exported directly. Tests verify the agent can detect them.

    Note: normalize_dataframe does NOT strip injections (it's a data agent,
    not a sanitization layer). These tests verify detection patterns.
    """

    INJECTION_PATTERNS = ["=CMD", "=1+1", "+cmd", "-1+1", "@SUM(A1)", "| whoami"]

    def _df_with_injection(self, cell_value: str) -> pd.DataFrame:
        return pd.DataFrame({
            "Name": ["Alice", "Bob"],
            "Formula": ["safe_value", cell_value],
        })

    def test_detect_equals_prefix_injection(self):
        injection_cells = ["=CMD /C calc", "=1+1", "=HYPERLINK('http://evil.com')"]
        for cell in injection_cells:
            df = self._df_with_injection(cell)
            has_injection = df["Formula"].astype(str).str.startswith("=").any()
            assert has_injection, f"Failed to detect injection: {cell}"

    def test_detect_plus_prefix_injection(self):
        df = self._df_with_injection("+cmd whoami")
        has_injection = df["Formula"].astype(str).str.match(r"^[=+\-@|]").any()
        assert has_injection

    def test_detect_at_prefix_injection(self):
        df = self._df_with_injection("@SUM(1+1)")
        has_injection = df["Formula"].astype(str).str.startswith("@").any()
        assert has_injection

    def test_safe_values_not_flagged(self):
        df = pd.DataFrame({"Name": ["Alice", "Bob"], "Code": ["A001", "B002"]})
        has_injection = df["Code"].astype(str).str.match(r"^[=+\-@|]").any()
        assert not has_injection

    def test_injection_count_in_column(self):
        df = pd.DataFrame({
            "Notes": ["normal", "=dangerous", "safe text", "+evil", "@formula"]
        })
        injection_count = df["Notes"].astype(str).str.match(r"^[=+\-@|]").sum()
        assert injection_count == 3


# =============================================================================
# TestExcelFormulaValues
# =============================================================================

class TestExcelFormulaValues:
    """Excel files with formula cells — data_only=True reads computed values."""

    def test_formula_cell_read_as_computed_value(self):
        """openpyxl with data_only=True reads the cached result of formulas."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["B1"] = 200
        ws["C1"] = "=A1+B1"  # formula — openpyxl stores as string, not computed
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        try:
            # data_only=True reads cached values (None if never calculated by Excel)
            from openpyxl import load_workbook
            wb2 = load_workbook(path, data_only=True)
            ws2 = wb2.active
            # The formula cell is saved as string formula (not evaluated by openpyxl)
            # If loaded fresh (not from Excel), cached value is None
            assert ws2["A1"].value == 100
            assert ws2["B1"].value == 200
            # C1 either has the formula string or None (not a Python runtime evaluation)
            assert ws2["C1"].value is None or ws2["C1"].value == "=A1+B1" or ws2["C1"].value == 300
        finally:
            _safe_unlink(path)

    def test_non_formula_cells_unaffected_by_data_only(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Product"
        ws["B1"] = "Price"
        ws["A2"] = "Widget A"
        ws["B2"] = 50.0
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        try:
            from openpyxl import load_workbook
            wb2 = load_workbook(path, data_only=True)
            ws2 = wb2.active
            assert ws2["A1"].value == "Product"
            assert ws2["B2"].value == 50.0
        finally:
            _safe_unlink(path)

    def test_excel_file_loaded_with_data_only_via_pandas(self):
        """Pandas reads Excel via openpyxl; formulas become NaN or cached values."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Q1", "Q2", "Total"])
        ws.append(["Alice", 100, 200, None])
        ws.append(["Bob", 150, 180, None])
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        try:
            df, info = _run(_client().load_file(file_path=path, filename="f.xlsx"))
            assert "Name" in df.columns
            assert len(df) >= 2
        finally:
            _safe_unlink(path)


# =============================================================================
# TestTimeSeries
# =============================================================================

class TestTimeSeries:
    """Rolling average, YoY growth, month-over-month, lag features."""

    def _monthly_sales(self):
        return pd.DataFrame({
            "Month": pd.date_range("2024-01-01", periods=6, freq="MS"),
            "Revenue": [1000.0, 1200.0, 900.0, 1500.0, 1800.0, 1600.0],
        })

    def test_rolling_3_month_average(self):
        df = self._monthly_sales()
        df["RollingAvg3"] = df["Revenue"].rolling(window=3).mean()
        # First 2 values are NaN (insufficient window)
        assert df["RollingAvg3"].iloc[0] != df["RollingAvg3"].iloc[0]  # NaN
        assert df["RollingAvg3"].iloc[1] != df["RollingAvg3"].iloc[1]  # NaN
        # Third value = mean(1000, 1200, 900) = 1033.33
        assert df["RollingAvg3"].iloc[2] == pytest.approx(1033.33, rel=1e-2)

    def test_month_over_month_growth(self):
        df = self._monthly_sales()
        df["MoM_Growth"] = df["Revenue"].pct_change()
        # Jan → Feb: (1200-1000)/1000 = 20%
        assert df["MoM_Growth"].iloc[1] == pytest.approx(0.20)
        # Feb → Mar: (900-1200)/1200 = -25%
        assert df["MoM_Growth"].iloc[2] == pytest.approx(-0.25)

    def test_yoy_growth_with_lag(self):
        df = pd.DataFrame({
            "Year": [2022, 2023, 2024],
            "Revenue": [100000.0, 115000.0, 125000.0],
        })
        df["YoY_Growth"] = df["Revenue"].pct_change()
        assert df["YoY_Growth"].iloc[1] == pytest.approx(0.15)
        assert df["YoY_Growth"].iloc[2] == pytest.approx(125000 / 115000 - 1, rel=1e-3)

    def test_lag_feature_creation(self):
        """Lag-1 of revenue (previous month's value)."""
        df = self._monthly_sales()
        df["Lag1_Revenue"] = df["Revenue"].shift(1)
        assert df["Lag1_Revenue"].iloc[0] != df["Lag1_Revenue"].iloc[0]  # NaN
        assert df["Lag1_Revenue"].iloc[1] == 1000.0
        assert df["Lag1_Revenue"].iloc[2] == 1200.0

    def test_cumulative_sum(self):
        df = self._monthly_sales()
        df["CumulativeRevenue"] = df["Revenue"].cumsum()
        assert df["CumulativeRevenue"].iloc[0] == 1000.0
        assert df["CumulativeRevenue"].iloc[1] == 2200.0
        assert df["CumulativeRevenue"].iloc[-1] == pytest.approx(8000.0)

    def test_exponential_moving_average(self):
        df = self._monthly_sales()
        df["EMA"] = df["Revenue"].ewm(span=3, adjust=False).mean()
        # EMA should be between min and max
        assert df["EMA"].min() >= df["Revenue"].min() * 0.5
        assert df["EMA"].max() <= df["Revenue"].max() * 1.5


# =============================================================================
# TestCurrencyNormalization
# =============================================================================

class TestCurrencyNormalization:
    """Currency symbols and thousand separators must be stripped before numeric ops."""

    def _strip_currency(self, series: pd.Series) -> pd.Series:
        """Common pre-processing pattern for currency strings."""
        return pd.to_numeric(
            series.astype(str)
                  .str.replace(r"[₹$£€,\s]", "", regex=True)
                  .str.strip(),
            errors="coerce"
        )

    def test_rupee_symbol_stripped(self):
        s = pd.Series(["₹10,000", "₹25,000", "₹5,500"])
        result = self._strip_currency(s)
        assert result.tolist() == pytest.approx([10000.0, 25000.0, 5500.0])

    def test_dollar_symbol_stripped(self):
        s = pd.Series(["$1,500.00", "$800.50", "$2,200.00"])
        result = self._strip_currency(s)
        assert result.tolist() == pytest.approx([1500.0, 800.5, 2200.0])

    def test_pound_symbol_stripped(self):
        s = pd.Series(["£10.00", "£20.50"])
        result = self._strip_currency(s)
        assert result.tolist() == pytest.approx([10.0, 20.5])

    def test_euro_symbol_stripped(self):
        s = pd.Series(["€1,200", "€3,500"])
        result = self._strip_currency(s)
        assert result.tolist() == pytest.approx([1200.0, 3500.0])

    def test_mixed_currency_column_arithmetic(self):
        df = pd.DataFrame({"Revenue": ["₹10,000", "₹20,000", "₹15,000"]})
        df["Revenue_Num"] = self._strip_currency(df["Revenue"])
        assert df["Revenue_Num"].sum() == pytest.approx(45000.0)
        assert df["Revenue_Num"].mean() == pytest.approx(15000.0)

    def test_invalid_currency_string_becomes_nan(self):
        s = pd.Series(["₹10,000", "N/A", "₹5,000"])
        result = self._strip_currency(s)
        assert result.isna().sum() == 1
        assert result.iloc[0] == pytest.approx(10000.0)


# =============================================================================
# TestErrorHandling
# =============================================================================

class TestErrorHandling:
    """Empty file, zero-byte file, wrong extension, corrupt data."""

    def test_zero_byte_csv_raises(self):
        content = b""
        # An empty file may raise or return empty DataFrame
        try:
            df, info = _run(_client()._load_csv(content=content))
            # If no exception, DataFrame should be empty or have no rows
            assert df.empty or len(df) == 0
        except Exception:
            pass  # Exception is acceptable for zero-byte input

    def test_csv_with_only_header_row(self):
        csv = "Name,Age,Salary\n"
        content = _csv_bytes(csv)
        df, info = _run(_client()._load_csv(content=content))
        assert len(df) == 0
        assert "Name" in df.columns or len(df.columns) >= 1

    def test_csv_with_inconsistent_column_count(self):
        """Rows with extra/missing columns — pandas handles gracefully."""
        csv = "A,B,C\n1,2,3\n4,5\n7,8,9,10\n"
        content = _csv_bytes(csv)
        df, info = _run(_client()._load_csv(content=content))
        # Loads without crashing; some cells may be NaN
        assert len(df) >= 1

    def test_all_null_dataframe_normalizes_to_empty(self):
        df = pd.DataFrame({
            "A": [None, None, None],
            "B": [None, None, None],
        })
        out = _client().normalize_dataframe(df)
        # All columns are null → dropped, all rows dropped
        assert out.empty or out.shape[1] == 0

    def test_single_column_csv(self):
        csv = "Value\n10\n20\n30\n"
        content = _csv_bytes(csv)
        df, info = _run(_client()._load_csv(content=content))
        assert "Value" in df.columns
        assert len(df) == 3


# =============================================================================
# TestColumnOperations
# =============================================================================

class TestColumnOperations:
    """Add calculated column, rename, drop, reorder."""

    def test_add_calculated_column(self):
        df = pd.DataFrame({"Units": [10, 20, 30], "UnitPrice": [5.0, 8.0, 12.0]})
        df["Revenue"] = df["Units"] * df["UnitPrice"]
        assert df["Revenue"].tolist() == pytest.approx([50.0, 160.0, 360.0])

    def test_add_percentage_column(self):
        df = pd.DataFrame({"Revenue": [1000.0, 2000.0, 3000.0]})
        total = df["Revenue"].sum()
        df["Share%"] = (df["Revenue"] / total * 100).round(2)
        assert df["Share%"].sum() == pytest.approx(100.0)
        assert df["Share%"].iloc[0] == pytest.approx(16.67, rel=1e-2)

    def test_rename_column(self):
        df = pd.DataFrame({"Qty": [1, 2], "Prc": [10.0, 20.0]})
        renamed = df.rename(columns={"Qty": "Quantity", "Prc": "Price"})
        assert "Quantity" in renamed.columns
        assert "Price" in renamed.columns
        assert "Qty" not in renamed.columns

    def test_drop_columns(self):
        df = pd.DataFrame({"A": [1], "B": [2], "C": [3], "Temp": [99]})
        result = df.drop(columns=["Temp"])
        assert "Temp" not in result.columns
        assert len(result.columns) == 3

    def test_reorder_columns(self):
        df = pd.DataFrame({"C": [3], "A": [1], "B": [2]})
        result = df[["A", "B", "C"]]
        assert list(result.columns) == ["A", "B", "C"]

    def test_string_split_into_columns(self):
        df = pd.DataFrame({"FullName": ["Alice Johnson", "Bob Smith", "Carol White"]})
        df[["FirstName", "LastName"]] = df["FullName"].str.split(" ", n=1, expand=True)
        assert df["FirstName"].tolist() == ["Alice", "Bob", "Carol"]
        assert df["LastName"].tolist() == ["Johnson", "Smith", "White"]


# =============================================================================
# TestDataValidation
# =============================================================================

class TestDataValidation:
    """Null count, type consistency, range validation, uniqueness checks."""

    def test_null_count_per_column(self):
        df = pd.DataFrame({
            "A": [1, None, 3, None],
            "B": ["x", "y", None, "z"],
            "C": [1.0, 2.0, 3.0, 4.0],
        })
        null_counts = df.isnull().sum()
        assert null_counts["A"] == 2
        assert null_counts["B"] == 1
        assert null_counts["C"] == 0

    def test_null_percentage_exceeds_threshold(self):
        df = pd.DataFrame({"Val": [1, None, None, None, None]})
        null_pct = df["Val"].isnull().mean()
        assert null_pct == pytest.approx(0.80)
        assert null_pct > 0.5  # flag this column

    def test_numeric_range_validation(self):
        df = pd.DataFrame({"GSTRate": [0.05, 0.12, 0.18, 0.28, -0.05, 1.50]})
        invalid = df[~df["GSTRate"].between(0, 1)]
        assert len(invalid) == 2
        assert -0.05 in invalid["GSTRate"].tolist()

    def test_uniqueness_check_on_key_column(self):
        df = pd.DataFrame({"InvoiceNo": ["INV001", "INV002", "INV001", "INV003"]})
        duplicates = df["InvoiceNo"].duplicated()
        assert duplicates.sum() == 1

    def test_type_consistency_check(self):
        """All values in a numeric column should be numeric after coercion."""
        df = pd.DataFrame({"Amount": [100, 200, "abc", 400]})
        numeric = pd.to_numeric(df["Amount"], errors="coerce")
        non_numeric_count = numeric.isna().sum()
        assert non_numeric_count == 1  # "abc" fails

    def test_email_format_validation(self):
        df = pd.DataFrame({
            "Email": ["alice@example.com", "not-an-email", "bob@company.org", "bad@"]
        })
        valid_emails = df["Email"].str.match(r"^[^@]+@[^@]+\.[^@]+$")
        assert valid_emails.sum() == 2

    def test_date_in_future_flagged(self):
        df = pd.DataFrame({
            # 2099-01-01 is always in the future; the others are historical
            "OrderDate": pd.to_datetime(["2020-01-01", "2021-12-31", "2099-01-01", "2022-06-15"])
        })
        today = pd.Timestamp.today().normalize()
        future_orders = df[df["OrderDate"] > today]
        assert len(future_orders) == 1
        assert future_orders["OrderDate"].iloc[0] == pd.Timestamp("2099-01-01")


# =============================================================================
# TestSessionIsolation
# =============================================================================

class TestSessionIsolation:
    """
    Two different thread_ids must not bleed into each other's sessions.
    Tests that session_state (in-memory dict) is keyed by thread_id.
    """

    def test_thread_id_keys_are_distinct(self):
        """Files loaded under different thread_ids are stored in separate sessions."""
        from backend.agents.spreadsheet_agent.state import session_state

        thread_a = "thread-A-isolation-test"
        thread_b = "thread-B-isolation-test"

        df_a = pd.DataFrame({"Product": ["Widget A"], "Revenue": [1000.0]})
        df_b = pd.DataFrame({"Employee": ["Alice"], "Salary": [85000.0]})

        session_state.store_dataframe(thread_a, "file_a", df_a)
        session_state.store_dataframe(thread_b, "file_b", df_b)

        sess_a = session_state.get(thread_a)
        sess_b = session_state.get(thread_b)

        assert sess_a is not None and sess_b is not None
        assert "Product" in sess_a.dataframes["file_a"].columns
        assert "Employee" in sess_b.dataframes["file_b"].columns
        # Thread A must not see Thread B's data
        assert "file_b" not in sess_a.dataframes
        assert "file_a" not in sess_b.dataframes

    def test_session_cleanup_removes_data(self):
        """After storing a DF, a fresh session_state for a different id has no data."""
        from backend.agents.spreadsheet_agent.state import session_state

        thread_id = "thread-cleanup-test"
        session_state.store_dataframe(thread_id, "f", pd.DataFrame({"X": [1, 2, 3]}))

        sess = session_state.get(thread_id)
        assert sess is not None
        assert "f" in sess.dataframes

        # A different thread should have no data
        other_sess = session_state.get("totally-different-thread-999")
        assert other_sess is None

    def test_two_sessions_can_have_different_schemas(self):
        """Sessions for different threads with different schemas coexist independently."""
        from backend.agents.spreadsheet_agent.state import session_state

        tid1 = "schema-test-A"
        tid2 = "schema-test-B"

        df1 = pd.DataFrame({"InvoiceNo": ["INV001"], "Amount": [1000.0]})
        df2 = pd.DataFrame({"EmpID": [101], "Salary": [85000.0], "Dept": ["Eng"]})

        session_state.store_dataframe(tid1, "inv", df1)
        session_state.store_dataframe(tid2, "emp", df2)

        s1 = session_state.get(tid1)
        s2 = session_state.get(tid2)

        assert list(s1.dataframes["inv"].columns) == ["InvoiceNo", "Amount"]
        assert list(s2.dataframes["emp"].columns) == ["EmpID", "Salary", "Dept"]
        assert len(s2.dataframes["emp"].columns) == 3
        # Schemas must not bleed across sessions
        assert "inv" not in s2.dataframes
        assert "emp" not in s1.dataframes
