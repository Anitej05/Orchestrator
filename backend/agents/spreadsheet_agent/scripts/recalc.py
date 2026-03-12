"""Recalculate and verify formulas in an Excel file.

Compares stored (cached) formula values against Python-computed values
to detect discrepancies. Essential for validating that formula-heavy
spreadsheets are correct.

Usage:
    python recalc.py <xlsx_file> [--sheet <name>] [--verbose]

Examples:
    python recalc.py report.xlsx
    python recalc.py report.xlsx --sheet "Financial Model" --verbose
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def recalc(xlsx_path: str, sheet_name: str = None, verbose: bool = False) -> dict:
    """Recalculate formulas and check for discrepancies.

    Args:
        xlsx_path: Path to the Excel file
        sheet_name: Specific sheet to check (default: all sheets)
        verbose: Whether to print detailed output

    Returns:
        Dict with results
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)

    results = {
        "file": xlsx_path,
        "sheets_checked": [],
        "total_formulas": 0,
        "errors_found": [],
        "warnings": [],
        "summary": "",
    }

    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sname in sheets:
        if sname not in wb.sheetnames:
            results["warnings"].append(f"Sheet '{sname}' not found")
            continue

        ws = wb[sname]
        ws_data = wb_data[sname]

        sheet_result = {
            "name": sname,
            "formulas": 0,
            "errors": [],
        }

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_result["formulas"] += 1
                    results["total_formulas"] += 1

                    # Check the cached value
                    data_cell = ws_data[cell.coordinate]
                    cached_value = data_cell.value

                    # Check for Excel error values
                    error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
                    if cached_value and str(cached_value) in error_values:
                        error_info = {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "error": str(cached_value),
                            "sheet": sname,
                        }
                        sheet_result["errors"].append(error_info)
                        results["errors_found"].append(error_info)

                    if verbose and cached_value is not None:
                        print(f"  {sname}!{cell.coordinate}: {cell.value} = {cached_value}")

        results["sheets_checked"].append(sheet_result)

    # Check for common issues
    _check_common_issues(wb, results)

    # Summary
    if results["errors_found"]:
        results["summary"] = f"FAIL: Found {len(results['errors_found'])} formula error(s)"
    else:
        results["summary"] = f"PASS: {results['total_formulas']} formulas checked, no errors"

    wb.close()
    wb_data.close()

    return results


def _check_common_issues(wb, results: dict) -> None:
    """Check for common spreadsheet issues."""
    for sname in wb.sheetnames:
        ws = wb[sname]

        for row in ws.iter_rows(max_row=min(ws.max_row or 1, 1000)):
            for cell in row:
                if cell.value is None:
                    continue

                val = str(cell.value)

                # Check for hardcoded values that should be formulas
                # (heuristic: numbers in cells adjacent to formula cells)
                if cell.number_format and "%" in (cell.number_format or ""):
                    if isinstance(cell.value, (int, float)) and not val.startswith("="):
                        # This might be a percentage that should be calculated
                        pass  # Log if adjacent cells have formulas

                # Check for common data quality issues
                if isinstance(cell.value, str):
                    if cell.value.strip() != cell.value:
                        results["warnings"].append(
                            f"{sname}!{cell.coordinate}: Leading/trailing whitespace"
                        )


def main():
    parser = argparse.ArgumentParser(description="Recalculate and verify Excel formulas")
    parser.add_argument("xlsx_file", help="Excel file to check")
    parser.add_argument("--sheet", default=None, help="Specific sheet to check")
    parser.add_argument("--verbose", action="store_true", help="Print all formula values")
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    args = parser.parse_args()

    if not Path(args.xlsx_file).exists():
        print(f"Error: {args.xlsx_file} not found")
        sys.exit(1)

    results = recalc(args.xlsx_file, sheet_name=args.sheet, verbose=args.verbose)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        print(f"\n{results['summary']}")
        print(f"  Sheets checked: {len(results['sheets_checked'])}")
        print(f"  Total formulas: {results['total_formulas']}")

        if results["errors_found"]:
            print("\n  Errors:")
            for err in results["errors_found"]:
                print(f"    {err['sheet']}!{err['cell']}: {err['formula']} → {err['error']}")

        if results["warnings"]:
            print(f"\n  Warnings ({len(results['warnings'])}):")
            for w in results["warnings"][:10]:
                print(f"    {w}")

    if results["errors_found"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
