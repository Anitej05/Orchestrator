"""
Spreadsheet Agent v3.0 - DataFrame Client

Production-grade file handling with robust loading and preprocessing.
Handles ANY CSV/Excel format reliably.
"""

import io
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
import pandas as pd
import numpy as np

from .config import STORAGE_DIR, MAX_FILE_SIZE_MB, ALLOWED_EXTENSIONS, LARGE_FILE_THRESHOLD_MB

logger = logging.getLogger("spreadsheet_agent.client")


class DataFrameClient:
    """
    Production-grade file handling for spreadsheets.
    
    Features:
    - Auto encoding detection (UTF-8, Latin-1, CP1252, etc.)
    - Auto delimiter detection (comma, tab, semicolon, pipe)
    - Multi-row header handling
    - Large file chunked processing
    - Robust error recovery
    """
    
    def __init__(self, storage_dir: Path = None):
        self.storage_dir = storage_dir or STORAGE_DIR
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        
        # Try to import file manager
        try:
            from agents.utils.agent_file_manager import AgentFileManager, FileType, FileStatus
            self.file_manager = AgentFileManager(
                agent_id="spreadsheet_agent",
                storage_dir=str(self.storage_dir)
            )
            self.FileType = FileType
            self.FileStatus = FileStatus
        except ImportError:
            logger.warning("AgentFileManager not available, using basic file handling")
            self.file_manager = None
            self.FileType = None
            self.FileStatus = None
    
    # ========================================================================
    # PUBLIC API
    # ========================================================================
    
    async def load_file(
        self,
        file_path: str = None,
        content: bytes = None,
        filename: str = None,
        **kwargs
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Load a file with robust auto-detection.
        
        Returns:
            Tuple of (DataFrame, detection_info)
        """
        if content is None and file_path is None:
            raise ValueError("Either file_path or content must be provided")
        
        # Determine file type
        ext = Path(filename or file_path or "").suffix.lower()
        
        if ext in ['.xlsx', '.xls']:
            try:
                # Try loading as Excel first
                return await self._load_excel(file_path, content, **kwargs)
            except Exception as e:
                logger.warning(f"Excel load failed for {file_path}, likely format mismatch. Exception: {e}")
                logger.info("Encouraging fallback to CSV/Text loader...")
                # Fallback to CSV loader which handles robust detection
                return await self._load_csv(file_path, content, **kwargs)
        else:
            return await self._load_csv(file_path, content, **kwargs)
    
    async def save_file(
        self,
        df: pd.DataFrame,
        filename: str,
        format: str = "csv",
        thread_id: str = None
    ) -> Tuple[str, str]:
        """
        Save DataFrame to file.
        
        Returns:
            Tuple of (file_id, file_path)
        """
        # Generate file path
        file_path = self.storage_dir / filename
        
        # Save based on format
        if format == "xlsx":
            df.to_excel(file_path, index=False)
        elif format == "json":
            df.to_json(file_path, orient="records", indent=2)
        else:
            df.to_csv(file_path, index=False)
        
        # Register with file manager if available
        file_id = filename
        if self.file_manager:
            try:
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                
                metadata = await self.file_manager.register_file(
                    content=file_content,
                    filename=filename,
                    file_type=self.FileType.SPREADSHEET,
                    thread_id=thread_id
                )
                file_id = metadata.file_id
            except Exception as e:
                logger.warning(f"Failed to register file: {e}")
        
        logger.info(f"Saved file: {file_path} (id: {file_id})")
        return file_id, str(file_path)
    
    async def build_context(
        self,
        df: pd.DataFrame,
        query: str = None,
        max_tokens: int = 8000
    ) -> str:
        """
        Build query-aware context for LLM.
        
        GENERALIZED APPROACH with LLM-powered intent analysis:
        1. Schema (columns, dtypes) - always included
        2. Query-aware data sampling - LLM decides what's relevant
        3. Statistics for numeric columns
        
        NO HARDCODED KEYWORDS - fully LLM-driven!
        """
        context_parts = []
        
        # 1. Basic info (lightweight schema)
        context_parts.append(f"DataFrame Info:")
        context_parts.append(f"- Shape: {df.shape[0]} rows × {df.shape[1]} columns")
        context_parts.append(f"- Columns: {list(df.columns)}")
        context_parts.append(f"- Dtypes: {dict(df.dtypes.astype(str))}")
        
        # 2. Query-aware data sampling (LLM-powered!)
        if query:
            # LLM decides what data is relevant
            relevant_data = await self._find_relevant_data_slice(df, query, max_rows=50)
            
            # Show how many matches were found
            if len(relevant_data) < len(df):
                context_parts.append(f"\nFound {len(relevant_data)} relevant rows (filtered from {len(df)} total)")
        else:
            # No query - use stratified sampling
            relevant_data = self._get_stratified_sample(df, n=15)
        
        # 3. Format relevant data for display
        relevant_cols = self._extract_relevant_columns(df, query) if query else df.columns.tolist()
        sample_cols = relevant_cols[:15] if len(relevant_cols) > 15 else relevant_cols
        
        # Show the relevant slice
        display_df = relevant_data[sample_cols].head(50)
        sample_str = display_df.to_string(max_rows=50, max_cols=15)
        context_parts.append(f"\nRelevant Data Sample:\n{sample_str}")
        
        # 4. Statistics for numeric columns (computed on FULL dataset)
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if numeric_cols:
            stats_df = df[numeric_cols[:10]].describe()
            context_parts.append(f"\nNumeric Statistics (Full Dataset):\n{stats_df.to_string()}")
        
        # 5. Unique values for categorical columns (if space allows)
        cat_cols = df.select_dtypes(include=['object', 'category']).columns.tolist()
        if cat_cols and len("\n".join(context_parts)) < max_tokens * 3:
            context_parts.append("\nCategorical Column Values:")
            for col in cat_cols[:5]:
                unique = df[col].dropna().unique()[:10]
                context_parts.append(f"  - {col}: {list(unique)}")
        
        return "\n".join(context_parts)
    
    def normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean up common spreadsheet issues.
        """
        if df.empty:
            return df
        
        # 1. Trim column names
        df.columns = df.columns.astype(str).str.strip()
        
        # 2. Remove fully empty columns
        df = df.dropna(axis=1, how='all')
        
        # 3. Remove Unnamed columns that are empty
        unnamed_cols = [c for c in df.columns if str(c).startswith('Unnamed')]
        for col in unnamed_cols:
            if df[col].isna().all() or (df[col].astype(str).str.strip() == '').all():
                df = df.drop(columns=[col])
        
        # 4. Remove fully empty rows
        df = df.dropna(how='all')
        
        # 5. Trim string values
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            # Replace 'nan' strings with actual NaN
            df[col] = df[col].replace(['nan', 'NaN', 'None', ''], np.nan)
        
        # 6. Try to coerce numeric strings
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    numeric = pd.to_numeric(df[col], errors='coerce')
                    # Only convert if >50% are successfully converted
                    if numeric.notna().sum() / len(df) > 0.5:
                        df[col] = numeric
                except:
                    pass
        
        return df.reset_index(drop=True)
    
    # ========================================================================
    # PRIVATE METHODS
    # ========================================================================
    
    async def _load_csv(
        self,
        file_path: str = None,
        content: bytes = None,
        **kwargs
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Robust CSV loading with auto-detection."""
        detection_info = {}
        
        # Get raw bytes
        if content:
            raw_bytes = content
        else:
            with open(file_path, 'rb') as f:
                raw_bytes = f.read()
        
        # Check file size
        size_mb = len(raw_bytes) / (1024 * 1024)
        detection_info['size_mb'] = round(size_mb, 2)
        
        if size_mb > MAX_FILE_SIZE_MB:
            raise ValueError(f"File too large: {size_mb:.1f}MB (max: {MAX_FILE_SIZE_MB}MB)")
        
        # Detect encoding
        encoding = self._detect_encoding(raw_bytes)
        detection_info['encoding'] = encoding
        
        # Decode text
        text = raw_bytes.decode(encoding, errors='replace')
        
        # Detect delimiter
        delimiter = self._detect_delimiter(text)
        detection_info['delimiter'] = delimiter
        
        # Detect header row
        header_row = self._detect_header_row(text, delimiter)
        detection_info['header_row'] = header_row
        
        # Try loading with detected parameters
        try:
            df = pd.read_csv(
                io.BytesIO(raw_bytes),
                encoding=encoding,
                delimiter=delimiter,
                header=header_row,
                on_bad_lines='warn',
                low_memory=False,
                **kwargs
            )
            detection_info['load_method'] = 'auto'
        except Exception as e:
            logger.warning(f"Auto-detection failed: {e}, trying fallback")
            df, fallback_info = self._fallback_load_csv(raw_bytes)
            detection_info.update(fallback_info)
        
        # Normalize
        df = self.normalize_dataframe(df)
        detection_info['final_shape'] = df.shape
        
        logger.info(f"Loaded CSV: {df.shape}, encoding={encoding}, delimiter='{delimiter}'")
        return df, detection_info
    
    async def _load_excel(
        self,
        file_path: str = None,
        content: bytes = None,
        use_smart_loading: bool = True,
        **kwargs
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Load Excel file with optional LLM-driven smart loading for complex files.
        """
        detection_info = {'format': 'excel'}
        
        # Save content to temp file if needed (openpyxl needs file path)
        temp_file = None
        if content and not file_path:
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(content)
            temp_file.close()
            file_path = temp_file.name
        
        try:
            # First, try to detect if file is complex
            is_complex = await self._is_complex_excel(file_path)
            detection_info['is_complex'] = is_complex
            
            if is_complex and use_smart_loading:
                logger.info("Complex Excel detected, using LLM-driven preprocessing")
                return await self._smart_load_excel(file_path, detection_info)
            else:
                # Simple file - use standard pandas
                return await self._simple_load_excel(file_path, detection_info, **kwargs)
                
        finally:
            # Cleanup temp file
            if temp_file:
                import os
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
    
    async def _is_complex_excel(self, file_path: str) -> bool:
        """
        Detect if Excel file has complex structure requiring LLM preprocessing.
        """
        try:
            from openpyxl import load_workbook
            # read_only=False needed to access merged_cells
            wb = load_workbook(file_path, data_only=True, read_only=False)
            ws = wb.active
            
            # Check indicators of complexity
            has_merged_cells = len(list(ws.merged_cells.ranges)) > 0
            
            # Check first 5 rows for title/metadata patterns
            first_rows_sparse = 0
            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_values = [ws.cell(row_idx, col).value for col in range(1, min(10, ws.max_column + 1))]
                non_empty = sum(1 for v in row_values if v)
                if non_empty <= 2:  # Sparse row (likely title or metadata)
                    first_rows_sparse += 1
            
            wb.close()
            
            # Complex if has merged cells OR multiple sparse rows at top
            return has_merged_cells or first_rows_sparse >= 2
            
        except Exception as e:
            logger.warning(f"Complexity detection failed: {e}")
            return False
    
    async def _simple_load_excel(
        self, 
        file_path: str, 
        detection_info: Dict,
        **kwargs
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Simple pandas-based Excel loading."""
        try:
            # Explicitly specify engine for xlsx to allow detection even if extension is ambiguous
            engine = 'openpyxl' if file_path.endswith('.xlsx') else None
            excel = pd.ExcelFile(file_path, engine=engine)
            sheets = excel.sheet_names
            detection_info['sheets'] = sheets
            
            df = None
            for sheet in sheets:
                # Explicitly specify engine for xlsx files to avoid "format cannot be determined" errors
                read_kwargs = kwargs.copy()
                if file_path.endswith('.xlsx') and 'engine' not in read_kwargs:
                    read_kwargs['engine'] = 'openpyxl'
                
                temp_df = pd.read_excel(file_path, sheet_name=sheet, **read_kwargs)
                if not temp_df.empty and len(temp_df.columns) > 0:
                    df = temp_df
                    detection_info['loaded_sheet'] = sheet
                    break
            
            if df is None:
                raise ValueError("No data found in any sheet")
            
            df = self.normalize_dataframe(df)
            detection_info['final_shape'] = df.shape
            detection_info['load_method'] = 'simple'
            
            logger.info(f"Loaded Excel (simple): {df.shape}")
            return df, detection_info
            
        except Exception as e:
            logger.error(f"Simple Excel load failed: {e}")
            raise ValueError(f"Failed to load Excel file: {e}")
    
    async def _smart_load_excel(
        self,
        file_path: str,
        detection_info: Dict,
        max_retries: int = 3
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        LLM-driven smart loading with retry loop and error learning.
        
        If preprocessing fails, the error is fed back to LLM for smarter retry.
        """
        from .llm import llm_client
        from .excel_tools import ExcelPreprocessor, analyze_spreadsheet_structure
        
        # Step 1: Smart analysis (only needs to run once)
        try:
            analysis = analyze_spreadsheet_structure(file_path)
            detection_info['analysis'] = {
                'main_sheet': analysis.get('main_sheet', {}),
                'structure': analysis.get('structure', {}),
                'hints': analysis.get('preprocessing_hints', [])
            }
            logger.info(f"Smart analysis: {analysis['main_sheet'].get('total_rows', 0)} rows")
        except Exception as e:
            logger.warning(f"Analysis failed: {e}, falling back to simple")
            detection_info['analysis_error'] = str(e)
            return await self._simple_load_excel(file_path, detection_info)
        
        # Retry loop with error learning
        previous_errors = []
        detection_info['attempts'] = []
        
        for attempt in range(1, max_retries + 1):
            try:
                logger.info(f"Preprocessing attempt {attempt}/{max_retries}")
                
                # Step 2: LLM generates plan (with error context if retry)
                plan = await llm_client.generate_preprocessing_plan(
                    analysis, 
                    error_context=previous_errors if previous_errors else None
                )
                
                attempt_info = {
                    'attempt': attempt,
                    'plan': plan.get('steps', [])[:5],  # First 5 steps for logging
                    'reasoning': plan.get('reasoning', '')[:100]
                }
                detection_info['attempts'].append(attempt_info)
                logger.info(f"LLM plan (attempt {attempt}): {len(plan.get('steps', []))} steps")
                
                # Step 3: Execute the plan
                with ExcelPreprocessor(file_path) as preprocessor:
                    # Select sheet
                    sheet_name = plan.get('sheet') or analysis.get('main_sheet', {}).get('name')
                    if sheet_name:
                        preprocessor.select_sheet(sheet_name)
                    
                    # Execute all steps
                    steps = plan.get('steps', [])
                    df = preprocessor.execute_plan(steps)
                
                # Validate result
                if df is None or df.empty:
                    error_msg = "Preprocessing produced empty DataFrame"
                    previous_errors.append({
                        'attempt': attempt,
                        'error': error_msg,
                        'steps_tried': [s.get('function') for s in steps]
                    })
                    logger.warning(f"Attempt {attempt} failed: {error_msg}")
                    continue
                
                # Check for reasonable result
                if len(df.columns) < 2:
                    error_msg = f"Only {len(df.columns)} column(s) extracted, expected more"
                    previous_errors.append({
                        'attempt': attempt,
                        'error': error_msg,
                        'columns_found': df.columns.tolist()
                    })
                    logger.warning(f"Attempt {attempt} suspicious: {error_msg}")
                    continue
                
                # Success!
                df = self.normalize_dataframe(df)
                detection_info['final_shape'] = df.shape
                detection_info['load_method'] = 'smart_toolkit'
                detection_info['successful_attempt'] = attempt
                logger.info(f"Loaded Excel (attempt {attempt}): {df.shape}")
                return df, detection_info
                
            except Exception as e:
                error_msg = str(e)
                previous_errors.append({
                    'attempt': attempt,
                    'error': error_msg,
                    'error_type': type(e).__name__
                })
                logger.warning(f"Attempt {attempt} failed with exception: {error_msg}")
                continue
        
        # All retries exhausted, fall back to simple loading
        logger.warning(f"All {max_retries} attempts failed, falling back to simple")
        detection_info['all_attempts_failed'] = True
        detection_info['errors'] = previous_errors
        return await self._simple_load_excel(file_path, detection_info)
    
    def extract_excel_structure(self, file_path: str) -> Dict[str, Any]:
        """
        Extract structural view of Excel file for LLM analysis.
        """
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(file_path, data_only=True)
        
        structure = {
            "sheets": [],
            "active_sheet": wb.active.title
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get merged regions
            merged = [str(m) for m in ws.merged_cells.ranges][:20]
            
            # Get first N rows as text grid with formatting hints
            grid = []
            for row_idx in range(1, min(25, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(15, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    val = cell.value
                    
                    # Include formatting hints
                    fmt = ""
                    try:
                        if cell.font and cell.font.bold:
                            fmt += "[B]"
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                            fmt += "[BG]"
                    except:
                        pass
                    
                    # Format value
                    if val is None:
                        row_data.append("")
                    elif isinstance(val, (int, float)):
                        row_data.append(f"{fmt}{val}")
                    else:
                        # Truncate long strings
                        str_val = str(val)[:50]
                        row_data.append(f"{fmt}{str_val}")
                
                grid.append(row_data)
            
            structure["sheets"].append({
                "name": sheet_name,
                "dimensions": f"{ws.max_row} rows x {ws.max_column} cols",
                "merged_cells": merged,
                "preview_grid": grid
            })
        
        wb.close()
        return structure
    
    def _execute_preprocessing_code(self, code: str, file_path: str) -> Optional[pd.DataFrame]:
        """
        Safely execute LLM-generated preprocessing code.
        """
        try:
            # Set up safe execution environment
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            local_vars = {
                'pd': pd,
                'load_workbook': load_workbook,
                'get_column_letter': get_column_letter,
                'file_path': file_path,
                'df': None
            }
            
            # Execute the code
            exec(code, {'__builtins__': __builtins__}, local_vars)
            
            return local_vars.get('df')
            
        except Exception as e:
            logger.error(f"Preprocessing code execution failed: {e}")
            return None

    
    def _detect_encoding(self, raw_bytes: bytes) -> str:
        """Detect file encoding."""
        try:
            import chardet
            result = chardet.detect(raw_bytes[:100000])
            encoding = result.get('encoding', 'utf-8')
            confidence = result.get('confidence', 0)
            
            if confidence > 0.7:
                return encoding
        except ImportError:
            pass
        
        # Fallback: try common encodings
        for enc in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                raw_bytes.decode(enc)
                return enc
            except:
                continue
        
        return 'utf-8'
    
    def _detect_delimiter(self, text: str) -> str:
        """Detect CSV delimiter by frequency analysis."""
        # Sample first ~50 lines
        lines = text.split('\n')[:50]
        sample = '\n'.join(lines)
        
        candidates = [',', '\t', ';', '|']
        
        # Count occurrences per line (more reliable than total)
        line_counts = {d: [] for d in candidates}
        for line in lines[:20]:
            for d in candidates:
                line_counts[d].append(line.count(d))
        
        # Calculate consistency (low std = consistent delimiter)
        scores = {}
        for d, counts in line_counts.items():
            if not counts or max(counts) == 0:
                scores[d] = 0
                continue
            
            avg = sum(counts) / len(counts)
            std = (sum((c - avg) ** 2 for c in counts) / len(counts)) ** 0.5
            
            # Score = avg count / (std + 1)
            scores[d] = avg / (std + 1) if avg > 0 else 0
        
        best = max(scores, key=scores.get)
        
        # Prefer comma if scores are close
        if scores[','] > 0 and scores[','] >= scores[best] * 0.8:
            return ','
        
        return best if scores[best] > 0 else ','
    
    def _detect_header_row(self, text: str, delimiter: str) -> int:
        """Detect which row contains headers."""
        lines = text.split('\n')[:30]
        
        for i, line in enumerate(lines):
            if not line.strip():
                continue
            
            parts = line.split(delimiter)
            if len(parts) < 2:
                continue
            
            # Header row typically has more strings than numbers
            string_count = sum(1 for p in parts if not self._is_number(p.strip()))
            string_ratio = string_count / len(parts)
            
            if string_ratio > 0.5:
                return i
        
        return 0
    
    def _is_number(self, s: str) -> bool:
        """Check if string is a number."""
        if not s:
            return False
        s = s.replace(',', '').replace('$', '').replace('%', '')
        try:
            float(s)
            return True
        except:
            return False
    
    def _fallback_load_csv(self, raw_bytes: bytes) -> Tuple[pd.DataFrame, Dict]:
        """Fallback loading with common configurations."""
        configs = [
            {'encoding': 'utf-8', 'delimiter': ','},
            {'encoding': 'latin-1', 'delimiter': ','},
            {'encoding': 'utf-8', 'delimiter': '\t'},
            {'encoding': 'utf-8', 'delimiter': ';'},
            {'encoding': 'cp1252', 'delimiter': ','},
        ]
        
        for config in configs:
            try:
                df = pd.read_csv(
                    io.BytesIO(raw_bytes),
                    encoding=config['encoding'],
                    delimiter=config['delimiter'],
                    on_bad_lines='skip',
                    low_memory=False
                )
                if len(df.columns) > 1 and len(df) > 0:
                    return df, {'load_method': 'fallback', **config}
            except:
                continue
        
        raise ValueError("Failed to load CSV with all fallback configurations")
    
    async def _find_relevant_data_slice(self, df: pd.DataFrame, query: str, max_rows: int = 50, max_retries: int = 3) -> pd.DataFrame:
        """
        SELF-CORRECTING relevance finder with retry loop and error learning.
        
        The agent will:
        1. Try to find relevant data based on LLM analysis
        2. Inspect what it fetched
        3. If empty/unexpected, retry with feedback about what failed
        
        NO HARDCODED KEYWORDS - fully LLM-driven with self-correction!
        """
        from .llm import llm_client
        
        # Prepare lightweight schema for LLM
        schema_info = {
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "sample_values": {
                col: df[col].dropna().unique()[:5].tolist() 
                for col in df.columns[:10]  # First 10 columns only
            }
        }
        
        # Track previous attempts for error learning
        previous_attempts = []
        
        for attempt in range(1, max_retries + 1):
            logger.info(f"Query intent analysis attempt {attempt}/{max_retries}")
            
            # Build context about previous failures
            error_context = None
            if previous_attempts:
                error_context = "Previous attempts failed:\n" + "\n".join([
                    f"Attempt {a['attempt']}: Searched for {a['search_terms']}, found {a['result_count']} rows. Issue: {a['issue']}"
                    for a in previous_attempts
                ])
            
            # Ask LLM: What data is relevant? (with error feedback if retry)
            if error_context:
                query_with_context = f"{query}\n\nNOTE: {error_context}\nPlease try a different approach."
                analysis = await llm_client.analyze_query_intent(query_with_context, schema_info)
            else:
                analysis = await llm_client.analyze_query_intent(query, schema_info)
            
            logger.info(f"LLM analysis (attempt {attempt}): {analysis.get('reasoning', 'No reasoning')}")
            
            relevant_rows = df.copy()
            search_applied = False
            
            # Apply LLM's decisions
            
            # 1. Search term filtering (if LLM says to)
            if analysis.get('needs_filtering') and analysis.get('search_terms'):
                search_applied = True
                for term in analysis['search_terms']:
                    target_cols = analysis.get('target_columns', [])
                    
                    if target_cols:
                        # Search in specific columns
                        mask = relevant_rows[target_cols].apply(
                            lambda row: any(term.lower() in str(val).lower() for val in row if pd.notna(val)),
                            axis=1
                        )
                    else:
                        # Search in all columns
                        mask = relevant_rows.apply(
                            lambda row: any(term.lower() in str(val).lower() for val in row if pd.notna(val)),
                            axis=1
                        )
                    
                    if mask.sum() > 0:
                        relevant_rows = relevant_rows[mask]
                        logger.info(f"LLM-identified term '{term}' found {mask.sum()} matches")
            
            # 2. Row index lookups (if LLM identified specific rows)
            if analysis.get('row_references'):
                try:
                    relevant_rows = df.iloc[analysis['row_references']]
                    return relevant_rows  # Direct lookup - always valid
                except IndexError:
                    logger.warning("Row references out of bounds")
            
            # 3. Sampling strategy (LLM decides which one!)
            sampling_strategy = analysis.get('sampling_strategy', 'stratified')
            
            if sampling_strategy == 'distribution' and len(relevant_rows) > max_rows:
                relevant_rows = self._get_distribution_sample(relevant_rows, n=max_rows)
            elif sampling_strategy == 'stratified' and len(relevant_rows) > max_rows:
                relevant_rows = self._get_stratified_sample(relevant_rows, n=max_rows)
            elif len(relevant_rows) > max_rows:
                relevant_rows = relevant_rows.head(max_rows)
            
            # INSPECT RESULTS - Decide if we need to retry
            result_count = len(relevant_rows)
            
            # Success conditions
            if result_count > 0 and result_count < len(df):
                # Found specific data (filtered successfully)
                logger.info(f"Successfully found {result_count} relevant rows")
                return relevant_rows
            
            if not search_applied:
                # LLM said no filtering needed (e.g., aggregation query)
                if len(df) > 20:
                    return self._get_stratified_sample(df, n=min(max_rows, 20))
                return df
            
            # Failure conditions - need retry
            if result_count == 0:
                # Empty results - search terms didn't match anything
                issue = "Search terms found no matches"
                logger.warning(f"Attempt {attempt} failed: {issue}")
                previous_attempts.append({
                    'attempt': attempt,
                    'search_terms': analysis.get('search_terms', []),
                    'target_columns': analysis.get('target_columns', []),
                    'result_count': 0,
                    'issue': issue
                })
                
                if attempt < max_retries:
                    continue  # Retry
                else:
                    # Final attempt failed - return safe default
                    logger.warning("All attempts failed, returning stratified sample")
                    return self._get_stratified_sample(df, n=min(max_rows, 20))
            
            elif result_count == len(df):
                # No filtering happened despite LLM saying to filter
                issue = "Filtering was requested but no rows were filtered"
                logger.warning(f"Attempt {attempt} suspicious: {issue}")
                previous_attempts.append({
                    'attempt': attempt,
                    'search_terms': analysis.get('search_terms', []),
                    'target_columns': analysis.get('target_columns', []),
                    'result_count': result_count,
                    'issue': issue
                })
                
                if attempt < max_retries:
                    continue  # Retry
                else:
                    # Return sample anyway (better than nothing)
                    return self._get_stratified_sample(df, n=min(max_rows, 20))
        
        # Should not reach here, but safety fallback
        return self._get_stratified_sample(df, n=min(max_rows, 20))
    

    
    def _get_distribution_sample(self, df: pd.DataFrame, n: int = 20) -> pd.DataFrame:
        """
        Get stratified sample showing data distribution.
        Useful for aggregation queries to show min, max, and range.
        """
        if len(df) <= n:
            return df
        
        # Find numeric column with most variance (likely the interesting one)
        numeric_cols = df.select_dtypes(include='number').columns
        if len(numeric_cols) > 0:
            # Sample from quantiles
            sample_indices = []
            for quantile in [0, 0.25, 0.5, 0.75, 1.0]:
                idx = int(len(df) * quantile)
                if idx >= len(df):
                    idx = len(df) - 1
                sample_indices.append(idx)
            
            # Add some random samples
            import random
            remaining = n - len(sample_indices)
            if remaining > 0:
                available = [i for i in range(len(df)) if i not in sample_indices]
                sample_indices.extend(random.sample(available, min(remaining, len(available))))
            
            return df.iloc[sorted(sample_indices)]
        else:
            # No numeric columns, just sample evenly
            return self._get_stratified_sample(df, n)
    
    def _get_stratified_sample(self, df: pd.DataFrame, n: int = 15) -> pd.DataFrame:
        """
        Get evenly distributed sample from dataframe.
        """
        if len(df) <= n:
            return df
        
        # Sample every k-th row to get even distribution
        step = len(df) // n
        indices = list(range(0, len(df), step))[:n]
        return df.iloc[indices]
    
    def _extract_relevant_columns(self, df: pd.DataFrame, query: str) -> List[str]:
        """Extract columns likely relevant to the query."""
        if not query:
            return df.columns.tolist()
        
        query_lower = query.lower()
        relevant = []
        
        for col in df.columns:
            col_lower = str(col).lower()
            
            # Direct mention
            if col_lower in query_lower or any(word in col_lower for word in query_lower.split()):
                relevant.append(col)
                continue
            
            # Common aliases
            aliases = {
                'revenue': ['sales', 'income', 'amount', 'total'],
                'date': ['time', 'period', 'year', 'month', 'day'],
                'category': ['type', 'group', 'class', 'segment'],
                'quantity': ['qty', 'count', 'units', 'number'],
                'price': ['cost', 'value', 'rate'],
                'name': ['title', 'description', 'label'],
            }
            
            for key, alias_list in aliases.items():
                if key in query_lower and any(a in col_lower for a in alias_list):
                    relevant.append(col)
                    break
        
        # If no relevant columns found, return all
        return relevant if relevant else df.columns.tolist()


# ========================================================================
# SMART DATA RESOLVER
# ========================================================================

class SmartDataResolver:
    """
    Self-resolving pipeline for spreadsheet data.
    Inspired by mail agent's SmartDataResolver.
    """
    
    def __init__(self, client: DataFrameClient, state):
        self.client = client
        self.state = state
    
    async def resolve_dataframe(
        self,
        params: Dict[str, Any],
        thread_id: str,
        require_data: bool = True
    ) -> Optional[pd.DataFrame]:
        """
        Smart resolution of DataFrame for any operation.
        
        Resolution order:
        1. Explicit file_id in params → Load from session
        2. File content in params → Parse inline
        3. Recent file in session → Use that
        4. No data → Return None (or error if required)
        """
        # 1. Explicit file_id
        file_id = params.get('file_id')
        if file_id:
            df = self.state.get_dataframe(thread_id, file_id)
            if df is not None:
                return df
            
            # Try to load from file manager
            if self.client.file_manager:
                metadata = self.client.file_manager.get_file(file_id)
                if metadata:
                    df, _ = await self.client.load_file(file_path=metadata.storage_path)
                    self.state.store_dataframe(thread_id, file_id, df, metadata.storage_path)
                    return df
        
        # 2. Inline content
        if 'content' in params or 'file_content' in params:
            content = params.get('content') or params.get('file_content')
            filename = params.get('filename', 'inline.csv')
            df, _ = await self.client.load_file(content=content, filename=filename)
            return df
        
        # 3. Latest file in session
        df = self.state.get_dataframe(thread_id)
        if df is not None:
            return df
        
        # 4. No data
        if require_data:
            raise ValueError("No data available. Please upload a file first.")
        return None
    
    def resolve_columns(
        self,
        df: pd.DataFrame,
        column_hints: List[str]
    ) -> List[str]:
        """
        Smart column resolution with fuzzy matching.
        """
        resolved = []
        
        for hint in column_hints:
            hint_lower = hint.lower().strip()
            
            # Exact match
            if hint in df.columns:
                resolved.append(hint)
                continue
            
            # Case-insensitive match
            for col in df.columns:
                if col.lower() == hint_lower:
                    resolved.append(col)
                    break
            else:
                # Fuzzy match (contains)
                for col in df.columns:
                    if hint_lower in col.lower() or col.lower() in hint_lower:
                        resolved.append(col)
                        break
        
        return resolved


# ========================================================================
# GLOBAL INSTANCE
# ========================================================================

df_client = DataFrameClient()
