"""
Spreadsheet Agent v3.0 - Excel Preprocessing Toolkit

Predefined functions for preprocessing complex Excel files.
The LLM analyzes files and decides which functions to call.
"""

import logging
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import pandas as pd
from datetime import datetime

logger = logging.getLogger("spreadsheet_agent.excel_tools")


# ============================================================================
# AVAILABLE TOOLS (For LLM prompt)
# ============================================================================

AVAILABLE_TOOLS = {
    "select_sheet": {
        "description": "Select which sheet to process",
        "params": {"sheet_name": "Name of the sheet to use"}
    },
    "unmerge_and_fill": {
        "description": "Handle merged cells by filling all cells with the top-left value. CALL THIS FIRST if file has merged cells.",
        "params": {"strategy": "'fill_all' (default) or 'first_value' (only keep first)"}
    },
    "skip_rows": {
        "description": "Skip metadata/title rows at the top (adjusts where data extraction starts)",
        "params": {"count": "Number of rows to skip from top"}
    },
    "set_header_row": {
        "description": "Define which row contains the column headers (1-indexed from top of file)",
        "params": {"row_number": "Row number (1-indexed) containing headers"}
    },
    "skip_columns": {
        "description": "Remove specific columns by name or letter (e.g., 'A', 'B') or 0-indexed position",
        "params": {"columns": "List of column names, letters, or 0-indexed positions to remove"}
    },
    "flatten_multi_headers": {
        "description": "Combine multiple header rows into a single header row by joining values",
        "params": {"row_count": "Number of header rows to combine (default: 2)", "separator": "String to join headers (default: ' - ')"}
    },
    "remove_empty_rows": {
        "description": "Remove rows that are completely empty (all cells are None/blank)",
        "params": {}
    },
    "remove_empty_columns": {
        "description": "Remove columns that are completely empty",
        "params": {}
    },
    "strip_whitespace": {
        "description": "Remove leading/trailing whitespace from all string values",
        "params": {}
    },
    "normalize_column_names": {
        "description": "Clean column names: lowercase, replace spaces/special chars with underscore",
        "params": {"lowercase": "Convert to lowercase (default: true)", "replace_spaces": "Replace spaces with underscore (default: true)"}
    },
    "detect_and_set_data_region": {
        "description": "Auto-detect where the actual data table begins by finding the first row with multiple values",
        "params": {}
    },
    "fill_merged_headers": {
        "description": "Fill empty header cells from left neighbor (for headers that span multiple columns)",
        "params": {}
    },
    "trim_trailing_empty_rows": {
        "description": "Remove empty rows at the bottom of the data (common in Excel exports)",
        "params": {}
    },
    "convert_dates": {
        "description": "Convert Excel serial dates to datetime objects in date-like columns",
        "params": {}
    },
    "remove_totals_row": {
        "description": "Remove the last row if it appears to be a totals/summary row",
        "params": {}
    }
}


def get_tools_prompt() -> str:
    """Generate tools documentation for LLM prompt."""
    lines = ["AVAILABLE PREPROCESSING FUNCTIONS:\n"]
    for name, info in AVAILABLE_TOOLS.items():
        params_str = ", ".join(f"{k}: {v}" for k, v in info["params"].items()) if info["params"] else "none"
        lines.append(f"- {name}: {info['description']}")
        lines.append(f"  Parameters: {params_str}\n")
    return "\n".join(lines)


# ============================================================================
# EXCEL PREPROCESSOR
# ============================================================================

class ExcelPreprocessor:
    """
    Toolkit for preprocessing complex Excel files.
    
    Usage:
        with ExcelPreprocessor(file_path) as preprocessor:
            df = preprocessor.execute_plan(steps)
    """
    
    def __init__(self, file_path: str):
        from openpyxl import load_workbook
        
        self.file_path = file_path
        # Note: data_only=True to get computed values instead of formulas
        # read_only=False because we may need to unmerge cells
        self.wb = load_workbook(file_path, data_only=True, read_only=False)
        self.ws = self.wb.active
        self.selected_sheet = self.ws.title
        
        # Configuration
        self.header_row = 1
        self.data_start_row = 2
        self.skip_top = 0
        self.columns_to_skip = []
        self.multi_header_rows = 1
        
        # Flags
        self._unmerged = False
        self._remove_empty_rows = True  # Default on
        self._remove_empty_cols = True  # Default on
        self._strip_whitespace = True   # Default on
        self._normalize_lowercase = False
        self._normalize_spaces = False
        self._convert_dates = False
        self._remove_totals = False
        
        # Cached data
        self._headers = None
        
        logger.info(f"ExcelPreprocessor initialized: {file_path}, sheet={self.selected_sheet}")
    
    def execute_plan(self, steps: List[Dict[str, Any]]) -> pd.DataFrame:
        """Execute a preprocessing plan (list of function calls)."""
        logger.info(f"Executing {len(steps)} preprocessing steps")
        
        for step in steps:
            func_name = step.get("function")
            params = step.get("params", {})
            
            if func_name == "convert_to_dataframe":
                continue  # Skip - we always call this at the end
            
            if hasattr(self, func_name):
                try:
                    logger.info(f"Step: {func_name}({params})")
                    getattr(self, func_name)(**params)
                except Exception as e:
                    logger.warning(f"Step {func_name} failed: {e}")
            else:
                logger.warning(f"Unknown function: {func_name}")
        
        # Always convert to DataFrame at the end
        return self.convert_to_dataframe()
    
    # ========================================================================
    # PREPROCESSING FUNCTIONS
    # ========================================================================
    
    def select_sheet(self, sheet_name: str):
        """Select which sheet to process."""
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
            self.selected_sheet = sheet_name
            self._headers = None  # Reset headers
            logger.info(f"Selected sheet: {sheet_name}")
        else:
            available = ", ".join(self.wb.sheetnames[:5])
            logger.warning(f"Sheet '{sheet_name}' not found. Available: {available}")
    
    def unmerge_and_fill(self, strategy: str = "fill_all"):
        """
        Handle merged cells by filling with top-left value.
        
        IMPORTANT: Must be called before other operations that depend on cell values.
        """
        if self._unmerged:
            logger.info("Already unmerged, skipping")
            return
        
        # Get list of merged ranges BEFORE modifying (important!)
        merged_ranges = list(self.ws.merged_cells.ranges)
        
        if not merged_ranges:
            logger.info("No merged cells found")
            self._unmerged = True
            return
        
        for merged_range in merged_ranges:
            # Get bounds: min_col, min_row, max_col, max_row
            min_col, min_row, max_col, max_row = merged_range.bounds
            
            # Get the value from top-left cell (the only cell with value)
            top_left_value = self.ws.cell(min_row, min_col).value
            
            # Unmerge the cells first
            self.ws.unmerge_cells(str(merged_range))
            
            # Fill values based on strategy
            if strategy == "fill_all":
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        self.ws.cell(row, col).value = top_left_value
            # "first_value" strategy: only the top-left cell keeps the value (already there)
        
        self._unmerged = True
        logger.info(f"Unmerged and filled {len(merged_ranges)} cell ranges")
    
    def skip_rows(self, count: int):
        """Skip N rows from the top when extracting data."""
        if count < 0:
            logger.warning(f"Invalid skip count: {count}, using 0")
            count = 0
        self.skip_top = count
        # Adjust header row if it would be skipped
        if self.header_row <= count:
            self.header_row = count + 1
            self.data_start_row = count + 2
        logger.info(f"Will skip {count} top rows, header now at row {self.header_row}")
    
    def set_header_row(self, row_number: int):
        """Set which row contains headers (1-indexed, relative to skip_top).
        
        If skip_rows was called first, this row_number is relative to the
        remaining rows after skipping.
        """
        if row_number < 1:
            logger.warning(f"Invalid header row: {row_number}, using 1")
            row_number = 1
        
        # Adjust for skip offset - the row_number is relative to remaining data
        actual_row = self.skip_top + row_number
        
        if actual_row > self.ws.max_row:
            logger.warning(f"Header row {actual_row} exceeds max row {self.ws.max_row}")
            actual_row = self.skip_top + 1
            
        self.header_row = actual_row
        self.data_start_row = actual_row + 1
        self._headers = None  # Reset cached headers
        logger.info(f"Header row set to {actual_row} (relative row {row_number} + skip {self.skip_top}), data starts at {self.data_start_row}")
    
    def skip_columns(self, columns: List):
        """Mark columns to skip during extraction."""
        self.columns_to_skip = columns
        logger.info(f"Will skip columns: {columns}")
    
    def flatten_multi_headers(self, row_count: int = 2, separator: str = " - "):
        """Combine multiple header rows into one."""
        if row_count < 1:
            row_count = 2
            
        self.multi_header_rows = row_count
        
        combined_headers = []
        for col in range(1, self.ws.max_column + 1):
            parts = []
            for row in range(self.header_row, self.header_row + row_count):
                if row <= self.ws.max_row:
                    val = self.ws.cell(row, col).value
                    if val is not None and str(val).strip():
                        parts.append(str(val).strip())
            
            if parts:
                combined_headers.append(separator.join(parts))
            else:
                combined_headers.append(f"Column_{col}")
        
        self._headers = combined_headers
        self.data_start_row = self.header_row + row_count
        logger.info(f"Flattened {row_count} header rows into {len(combined_headers)} columns")
    
    def remove_empty_rows(self):
        """Enable removal of completely empty rows."""
        self._remove_empty_rows = True
    
    def remove_empty_columns(self):
        """Enable removal of completely empty columns."""
        self._remove_empty_cols = True
    
    def strip_whitespace(self):
        """Enable whitespace stripping from string values."""
        self._strip_whitespace = True
    
    def normalize_column_names(self, lowercase: bool = True, replace_spaces: bool = True):
        """Enable column name normalization."""
        self._normalize_lowercase = lowercase
        self._normalize_spaces = replace_spaces
    
    def detect_and_set_data_region(self):
        """Auto-detect where the data table starts."""
        # Look for the first row with multiple non-empty cells that looks like headers
        for row_idx in range(1, min(50, self.ws.max_row + 1)):
            row_values = []
            for col in range(1, min(20, self.ws.max_column + 1)):
                val = self.ws.cell(row_idx, col).value
                row_values.append(val)
            
            non_empty = [v for v in row_values if v is not None and str(v).strip()]
            
            # Header row typically has 3+ values and most are strings
            if len(non_empty) >= 3:
                string_count = sum(1 for v in non_empty if isinstance(v, str))
                if string_count / len(non_empty) > 0.5:
                    self.header_row = row_idx
                    self.data_start_row = row_idx + 1
                    logger.info(f"Auto-detected header at row {row_idx}")
                    return
        
        logger.info("Auto-detection: using defaults (header row 1)")
    
    def fill_merged_headers(self):
        """Fill empty header cells from left neighbor."""
        if self._headers is None:
            self._extract_headers()
        
        filled = []
        last_value = "Column"
        
        for i, h in enumerate(self._headers):
            if h and str(h).strip() and not str(h).startswith("Column_"):
                last_value = str(h).strip()
                filled.append(last_value)
            else:
                # Use last value + suffix to ensure uniqueness
                filled.append(f"{last_value}_{i+1}")
        
        self._headers = filled
        logger.info("Filled empty headers from left neighbors")
    
    def trim_trailing_empty_rows(self):
        """Remove empty rows at the bottom (handled in convert_to_dataframe)."""
        pass  # Handled in conversion
    
    def convert_dates(self):
        """Enable date conversion for date-like columns."""
        self._convert_dates = True
    
    def remove_totals_row(self):
        """Enable removal of totals row if detected."""
        self._remove_totals = True
    
    def convert_to_dataframe(self) -> pd.DataFrame:
        """Convert processed worksheet to DataFrame."""
        # Extract headers if not already done
        if self._headers is None:
            self._extract_headers()
        
        # Determine actual number of columns
        num_cols = len(self._headers)
        
        # Extract data rows
        data = []
        for row_idx in range(self.data_start_row, self.ws.max_row + 1):
            row_data = []
            for col_idx in range(1, num_cols + 1):
                val = self.ws.cell(row_idx, col_idx).value
                
                # Strip whitespace from strings
                if self._strip_whitespace and isinstance(val, str):
                    val = val.strip()
                    if val == "":
                        val = None
                
                row_data.append(val)
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=self._headers)
        
        # Remove completely empty rows
        if self._remove_empty_rows:
            original_len = len(df)
            df = df.dropna(how='all')
            if len(df) < original_len:
                logger.info(f"Removed {original_len - len(df)} empty rows")
        
        # Remove completely empty columns
        if self._remove_empty_cols:
            empty_cols = df.columns[df.isna().all()].tolist()
            if empty_cols:
                df = df.drop(columns=empty_cols)
                logger.info(f"Removed {len(empty_cols)} empty columns")
        
        # Remove totals row if requested
        if self._remove_totals and len(df) > 1:
            last_row = df.iloc[-1]
            # Check if last row looks like totals (has "total" somewhere or is summary)
            last_row_str = " ".join(str(v).lower() for v in last_row.values if v is not None)
            if any(keyword in last_row_str for keyword in ['total', 'sum', 'grand', 'subtotal']):
                df = df.iloc[:-1]
                logger.info("Removed totals row")
        
        # Skip specific columns
        if self.columns_to_skip:
            cols_to_remove = []
            for col in self.columns_to_skip:
                if col in df.columns:
                    cols_to_remove.append(col)
                elif isinstance(col, int) and 0 <= col < len(df.columns):
                    cols_to_remove.append(df.columns[col])
                elif isinstance(col, str) and len(col) <= 2:
                    # Might be a column letter like 'A', 'B', 'AA'
                    col_idx = self._column_letter_to_index(col)
                    if col_idx is not None and col_idx < len(df.columns):
                        cols_to_remove.append(df.columns[col_idx])
            
            if cols_to_remove:
                df = df.drop(columns=cols_to_remove, errors='ignore')
                logger.info(f"Skipped columns: {cols_to_remove}")
        
        # Normalize column names
        if self._normalize_lowercase:
            df.columns = df.columns.astype(str).str.lower()
        if self._normalize_spaces:
            df.columns = df.columns.astype(str).str.replace(r'[^\w]', '_', regex=True)
            df.columns = df.columns.str.replace(r'_+', '_', regex=True)
            df.columns = df.columns.str.strip('_')
        
        # Remove Unnamed/Column_ columns that are empty
        unnamed = [c for c in df.columns if str(c).startswith('Column_') or str(c).startswith('Unnamed')]
        for col in unnamed:
            if df[col].isna().all() or (df[col].astype(str).str.strip() == '').all():
                df = df.drop(columns=[col])
        
        # Ensure unique column names
        df = self._ensure_unique_columns(df)
        
        logger.info(f"Converted to DataFrame: {df.shape}")
        return df.reset_index(drop=True)
    
    def _extract_headers(self):
        """Extract headers from the worksheet."""
        self._headers = []
        
        for col_idx in range(1, self.ws.max_column + 1):
            val = self.ws.cell(self.header_row, col_idx).value
            
            if val is not None:
                header = str(val).strip()
                if header:
                    self._headers.append(header)
                else:
                    self._headers.append(f"Column_{col_idx}")
            else:
                self._headers.append(f"Column_{col_idx}")
    
    def _column_letter_to_index(self, letter: str) -> Optional[int]:
        """Convert Excel column letter to 0-based index."""
        try:
            result = 0
            for char in letter.upper():
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result - 1
        except:
            return None
    
    def _ensure_unique_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ensure all column names are unique by adding suffixes."""
        cols = df.columns.tolist()
        seen = {}
        new_cols = []
        
        for col in cols:
            col_str = str(col)
            if col_str in seen:
                seen[col_str] += 1
                new_cols.append(f"{col_str}_{seen[col_str]}")
            else:
                seen[col_str] = 0
                new_cols.append(col_str)
        
        df.columns = new_cols
        return df
    
    def close(self):
        """Close workbook."""
        try:
            self.wb.close()
        except:
            pass
    
    def __enter__(self):
        return self
    
    def __exit__(self, *args):
        self.close()


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def preprocess_excel(
    file_path: str,
    steps: List[Dict[str, Any]]
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Preprocess an Excel file using a plan.
    
    Args:
        file_path: Path to Excel file
        steps: List of {"function": "name", "params": {...}}
    
    Returns:
        Tuple of (DataFrame, processing_info)
    """
    info = {
        "file": file_path,
        "steps_executed": len(steps),
        "timestamp": datetime.now().isoformat()
    }
    
    try:
        with ExcelPreprocessor(file_path) as preprocessor:
            df = preprocessor.execute_plan(steps)
            info["final_shape"] = df.shape
            info["columns"] = df.columns.tolist()
            info["success"] = True
    except Exception as e:
        logger.error(f"Preprocessing failed: {e}")
        info["success"] = False
        info["error"] = str(e)
        df = pd.DataFrame()
    
    return df, info


# ============================================================================
# SMART SPREADSHEET ANALYSIS
# ============================================================================

def analyze_spreadsheet_structure(file_path: str) -> Dict[str, Any]:
    """
    Human-like "glancing" at a spreadsheet to understand its structure.
    
    Works for files of any size by using surgical sampling instead of
    reading the entire file.
    
    Returns:
        Structured analysis suitable for LLM preprocessing decisions.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import random
    
    logger.info(f"Analyzing spreadsheet structure: {file_path}")
    
    analysis = {
        "file_info": {},
        "sheets": [],
        "main_sheet": {},
        "structure": {},
        "samples": {},
        "columns": [],
        "preprocessing_hints": []
    }
    
    try:
        # Note: read_only=False needed because some files don't have cached dimensions
        wb = load_workbook(file_path, data_only=True, read_only=False)
        
        # ================================================================
        # STEP 1: Quick Metadata Scan
        # ================================================================
        analysis["file_info"] = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames[:10]  # First 10 sheets
        }
        
        # Analyze each sheet briefly
        sheets_info = []
        main_sheet_name = None
        max_data = 0
        
        for sheet_name in wb.sheetnames[:5]:  # First 5 sheets
            ws = wb[sheet_name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            
            sheet_info = {
                "name": sheet_name,
                "rows": rows,
                "cols": cols,
                "has_data": rows > 1 and cols > 1
            }
            sheets_info.append(sheet_info)
            
            # Track sheet with most data
            if rows * cols > max_data and rows > 1:
                max_data = rows * cols
                main_sheet_name = sheet_name
        
        analysis["sheets"] = sheets_info
        
        # Use main sheet for detailed analysis
        if main_sheet_name is None:
            main_sheet_name = wb.sheetnames[0]
        
        ws = wb[main_sheet_name]
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        
        analysis["main_sheet"] = {
            "name": main_sheet_name,
            "total_rows": total_rows,
            "total_cols": total_cols
        }
        
        # ================================================================
        # STEP 2: Detailed Analysis (merged cells, etc.)
        # ================================================================
        # Already have ws from above, no need to reopen
        
        # Get merged cell info
        merged_cells = []
        for merged_range in list(ws.merged_cells.ranges)[:20]:
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_cells.append({
                "range": str(merged_range),
                "top_left": (min_row, min_col),
                "size": (max_row - min_row + 1, max_col - min_col + 1)
            })
        
        analysis["structure"]["merged_cells"] = merged_cells
        analysis["structure"]["has_merged"] = len(merged_cells) > 0
        
        # ================================================================
        # STEP 3: Strategic Row Sampling
        # ================================================================
        samples = {
            "top_rows": [],      # Title, metadata, headers
            "first_data": [],    # First few data rows
            "middle_sample": [], # Random middle rows
            "last_rows": []      # End rows (totals?)
        }
        
        # Top 10 rows (captures title, metadata, headers)
        for row_idx in range(1, min(11, total_rows + 1)):
            row_data = []
            for col_idx in range(1, min(15, total_cols + 1)):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                
                # Include formatting hints
                fmt = ""
                try:
                    if cell.font and cell.font.bold:
                        fmt = "[B]"
                    if cell.fill and cell.fill.fgColor and getattr(cell.fill.fgColor, 'rgb', None):
                        if cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                            fmt = "[BG]" + fmt
                except:
                    pass
                
                if val is None:
                    row_data.append("")
                else:
                    val_str = str(val)[:30]  # Truncate long values
                    row_data.append(f"{fmt}{val_str}" if fmt else val_str)
            
            samples["top_rows"].append(row_data)
        
        # First 3 data rows (after likely header)
        header_row = _detect_header_row(samples["top_rows"])
        data_start = header_row + 1
        
        for row_idx in range(data_start, min(data_start + 3, total_rows + 1)):
            row_data = []
            for col_idx in range(1, min(15, total_cols + 1)):
                val = ws.cell(row_idx, col_idx).value
                row_data.append(str(val)[:30] if val is not None else "")
            samples["first_data"].append(row_data)
        
        # 3 random middle rows (if file is large enough)
        if total_rows > 20:
            middle_start = max(10, total_rows // 4)
            middle_end = min(total_rows - 5, 3 * total_rows // 4)
            
            if middle_end > middle_start:
                random_rows = random.sample(range(middle_start, middle_end), min(3, middle_end - middle_start))
                for row_idx in sorted(random_rows):
                    row_data = []
                    for col_idx in range(1, min(15, total_cols + 1)):
                        val = ws.cell(row_idx, col_idx).value
                        row_data.append(str(val)[:30] if val is not None else "")
                    samples["middle_sample"].append(row_data)
        
        # Last 5 rows (to detect totals)
        for row_idx in range(max(1, total_rows - 4), total_rows + 1):
            row_data = []
            for col_idx in range(1, min(15, total_cols + 1)):
                val = ws.cell(row_idx, col_idx).value
                row_data.append(str(val)[:30] if val is not None else "")
            samples["last_rows"].append(row_data)
        
        analysis["samples"] = samples
        
        # ================================================================
        # STEP 4: Column Analysis
        # ================================================================
        columns_info = []
        
        # Get headers from detected header row
        headers = samples["top_rows"][header_row - 1] if header_row <= len(samples["top_rows"]) else []
        
        # Analyze each column (first 15)
        for col_idx in range(min(15, total_cols)):
            col_info = {
                "index": col_idx,
                "letter": get_column_letter(col_idx + 1),
                "header": headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx+1}"
            }
            
            # Sample values to determine type
            sample_values = []
            for row in samples["first_data"] + samples["middle_sample"]:
                if col_idx < len(row) and row[col_idx]:
                    sample_values.append(row[col_idx])
            
            col_info["inferred_type"] = _infer_column_type(sample_values)
            col_info["sample_values"] = sample_values[:3]
            columns_info.append(col_info)
        
        analysis["columns"] = columns_info
        
        # ================================================================
        # STEP 5: Structure Detection & Hints
        # ================================================================
        analysis["structure"]["detected_header_row"] = header_row
        analysis["structure"]["data_starts_row"] = data_start
        analysis["structure"]["header_confidence"] = _calculate_header_confidence(samples["top_rows"], header_row)
        
        # Generate preprocessing hints
        hints = []
        
        if analysis["structure"]["has_merged"]:
            hints.append("MERGED_CELLS: Call unmerge_and_fill() first")
        
        if header_row > 1:
            hints.append(f"TITLE_ROWS: {header_row - 1} rows before header, use skip_rows or set_header_row")
        
        # Check for multi-row headers
        if header_row <= len(samples["top_rows"]) - 1:
            hdr_row = samples["top_rows"][header_row - 1]
            next_row = samples["top_rows"][header_row] if header_row < len(samples["top_rows"]) else []
            if _looks_like_header(next_row) and not _looks_like_data(next_row):
                hints.append("MULTI_HEADER: May have multi-row headers, consider flatten_multi_headers")
        
        # Check for totals row
        if samples["last_rows"]:
            last_row_str = " ".join(samples["last_rows"][-1]).lower()
            if any(kw in last_row_str for kw in ["total", "sum", "grand", "subtotal"]):
                hints.append("TOTALS_ROW: Last row appears to be totals, use remove_totals_row")
        
        analysis["preprocessing_hints"] = hints
        
        wb.close()
        logger.info(f"Analysis complete: {total_rows} rows, {total_cols} cols, {len(hints)} hints")
        
    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        analysis["error"] = str(e)
    
    return analysis


def _detect_header_row(top_rows: List[List[str]]) -> int:
    """Detect which row is the header based on content patterns."""
    best_row = 1
    best_score = 0
    
    for idx, row in enumerate(top_rows):
        score = 0
        
        # Strip formatting markers to get actual content
        cleaned = []
        bold_count = 0
        bg_count = 0
        
        for v in row:
            if not v:
                continue
            
            # Track formatting markers
            has_bold = "[B]" in v
            has_bg = "[BG]" in v
            
            if has_bold:
                bold_count += 1
            if has_bg:
                bg_count += 1
            
            # Get the actual content without markers
            clean_v = v.replace("[B]", "").replace("[BG]", "").strip()
            if clean_v:
                cleaned.append(clean_v)
        
        # Skip rows with too few actual values
        if len(cleaned) < 2:
            continue
        
        # More non-empty cells = higher score
        score += len(cleaned) * 2
        
        # Bold cells WITH CONTENT = very likely header
        # (Row 5 has bold empty cells, Row 6 has bold headers)
        score += bold_count * 3
        
        # Background color = likely header
        score += bg_count * 2
        
        # Strings (not numbers/dates) = likely header
        string_count = sum(1 for v in cleaned if not _is_numeric(v) and not _looks_like_date(v))
        score += string_count * 2
        
        # Penalize rows that look like data (have dates, numbers)
        date_count = sum(1 for v in cleaned if _looks_like_date(v))
        numeric_count = sum(1 for v in cleaned if _is_numeric(v))
        score -= (date_count + numeric_count) * 2
        
        if score > best_score:
            best_score = score
            best_row = idx + 1
    
    return best_row


def _looks_like_date(value: str) -> bool:
    """Check if a string looks like a date."""
    if not value:
        return False
    # Common date patterns
    import re
    date_patterns = [
        r'\d{4}-\d{2}-\d{2}',  # 2025-03-31
        r'\d{2}/\d{2}/\d{4}',  # 03/31/2025
        r'\d{2}-\d{2}-\d{4}',  # 31-03-2025
    ]
    for pattern in date_patterns:
        if re.search(pattern, value):
            return True
    return False


def _looks_like_header(row: List[str]) -> bool:
    """Check if a row looks like a header."""
    non_empty = [v for v in row if v and v.strip()]
    if len(non_empty) < 2:
        return False
    
    # Headers are usually strings
    numeric_count = sum(1 for v in non_empty if _is_numeric(v))
    return numeric_count / len(non_empty) < 0.3


def _looks_like_data(row: List[str]) -> bool:
    """Check if a row looks like data (has numbers)."""
    non_empty = [v for v in row if v and v.strip()]
    if len(non_empty) < 2:
        return False
    
    numeric_count = sum(1 for v in non_empty if _is_numeric(v))
    return numeric_count / len(non_empty) > 0.3


def _is_numeric(value: str) -> bool:
    """Check if a string represents a number."""
    if not value:
        return False
    clean = value.replace(",", "").replace("$", "").replace("%", "").replace(" ", "")
    try:
        float(clean)
        return True
    except:
        return False


def _infer_column_type(values: List[str]) -> str:
    """Infer column data type from sample values."""
    if not values:
        return "unknown"
    
    numeric_count = 0
    date_count = 0
    
    date_patterns = ["date", "/", "-"]
    
    for v in values:
        if _is_numeric(v):
            numeric_count += 1
        elif any(p in v.lower() for p in date_patterns):
            date_count += 1
    
    total = len(values)
    if numeric_count / total > 0.7:
        return "numeric"
    elif date_count / total > 0.5:
        return "date"
    else:
        return "text"


def _calculate_header_confidence(top_rows: List[List[str]], header_row: int) -> float:
    """Calculate confidence score for header detection."""
    if header_row > len(top_rows):
        return 0.5
    
    row = top_rows[header_row - 1]
    non_empty = [v for v in row if v and v.strip()]
    
    if len(non_empty) < 2:
        return 0.3
    
    # Factors that increase confidence
    confidence = 0.5
    
    # Bold formatting
    if any("[B]" in v for v in row):
        confidence += 0.2
    
    # Background color
    if any("[BG]" in v for v in row):
        confidence += 0.1
    
    # All strings (no numbers)
    if not any(_is_numeric(v.replace("[B]", "").replace("[BG]", "")) for v in non_empty):
        confidence += 0.1
    
    # Next row looks like data
    if header_row < len(top_rows):
        next_row = top_rows[header_row]
        if _looks_like_data(next_row):
            confidence += 0.1
    
    return min(confidence, 1.0)

