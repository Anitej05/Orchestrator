"""
Advanced Edge Case Handler for Spreadsheet Agent

Handles industry-standard edge cases including:
- Merged cells
- Formula value extraction
- Error cell handling (#DIV/0!, #N/A, etc.)
- Inconsistent column counts
- Rich formatting interpretation

Requirements: 6.1-6.6, 15.1-15.7
"""

import logging
import re
from typing import Dict, Any, Optional, List, Tuple, Union
from pathlib import Path
import pandas as pd
import numpy as np

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available - advanced Excel features disabled")

logger = logging.getLogger(__name__)


# ============================================================================
# DATA MODELS
# ============================================================================

class MergedCellInfo:
    """Information about a merged cell region"""
    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int, value: Any):
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col
        self.value = value
        self.span_rows = max_row - min_row + 1
        self.span_cols = max_col - min_col + 1
    
    def __repr__(self):
        return f"MergedCell({self.min_row}:{self.max_row}, {self.min_col}:{self.max_col}, '{self.value}')"


class FormulaInfo:
    """Information about a formula cell"""
    def __init__(self, formula: str, calculated_value: Any, cell_address: str):
        self.formula = formula
        self.calculated_value = calculated_value
        self.cell_address = cell_address
        self.is_error = self._is_error_value(calculated_value)
    
    def _is_error_value(self, value: Any) -> bool:
        """Check if value is an Excel error"""
        if isinstance(value, str):
            return value.startswith('#') and value in [
                '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!'
            ]
        return False
    
    def __repr__(self):
        return f"Formula('{self.formula}' -> {self.calculated_value})"


# ============================================================================
# EDGE CASE HANDLER
# ============================================================================

class EdgeCaseHandler:
    """
    Handles advanced spreadsheet edge cases for robust parsing.
    
    Features:
    - Merged cell detection and unmerging
    - Formula value extraction
    - Error cell handling
    - Inconsistent column count normalization
    - Rich formatting interpretation
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.EdgeCaseHandler")
        self._merged_cells_cache = {}
        self._formulas_cache = {}
    
    def process_excel_file(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        handle_merged_cells: bool = True,
        extract_formulas: bool = True,
        handle_errors: bool = True
    ) -> pd.DataFrame:
        """
        Process Excel file with advanced edge case handling.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet to process (None for first sheet)
            handle_merged_cells: Whether to unmerge and replicate values
            extract_formulas: Whether to extract calculated values from formulas
            handle_errors: Whether to handle Excel error values
            
        Returns:
            Processed DataFrame
            
        Requirements: 6.1, 6.4, 6.5, 15.1, 15.2, 15.6
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl not available - falling back to basic pandas loading")
            return pd.read_excel(file_path, sheet_name=sheet_name or 0)
        
        try:
            # Load workbook with openpyxl for advanced features
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            # Select worksheet
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                else:
                    self.logger.warning(f"Sheet '{sheet_name}' not found, using first sheet")
                    worksheet = workbook.active
            else:
                worksheet = workbook.active
            
            self.logger.info(f"Processing Excel sheet: {worksheet.title}")
            
            # Extract merged cells info
            merged_cells = []
            if handle_merged_cells:
                merged_cells = self._extract_merged_cells(worksheet)
                self.logger.info(f"Found {len(merged_cells)} merged cell regions")
            
            # Extract formulas info
            formulas = []
            if extract_formulas:
                formulas = self._extract_formulas(worksheet)
                self.logger.info(f"Found {len(formulas)} formula cells")
            
            # Convert to DataFrame with edge case handling
            df = self._worksheet_to_dataframe(
                worksheet,
                merged_cells=merged_cells,
                formulas=formulas,
                handle_errors=handle_errors
            )
            
            # Handle inconsistent column counts
            df = self._normalize_column_counts(df)
            
            self.logger.info(f"Processed Excel file: {len(df)} rows × {len(df.columns)} cols")
            return df
            
        except Exception as e:
            self.logger.error(f"Advanced Excel processing failed: {e}, falling back to pandas")
            return pd.read_excel(file_path, sheet_name=sheet_name or 0)
    
    def handle_merged_cells_in_dataframe(
        self,
        df: pd.DataFrame,
        merged_regions: List[Tuple[int, int, int, int, Any]]
    ) -> pd.DataFrame:
        """
        Handle merged cells in an existing DataFrame.
        
        Args:
            df: Source DataFrame
            merged_regions: List of (min_row, max_row, min_col, max_col, value) tuples
            
        Returns:
            DataFrame with merged cells handled
            
        Requirements: 6.2
        """
        result_df = df.copy()
        
        for min_row, max_row, min_col, max_col, value in merged_regions:
            try:
                # Adjust for 0-based indexing
                start_row = max(0, min_row - 1)
                end_row = min(len(result_df), max_row)
                start_col = max(0, min_col - 1)
                end_col = min(len(result_df.columns), max_col)
                
                # Replicate value across merged region
                for row in range(start_row, end_row):
                    for col in range(start_col, end_col):
                        if col < len(result_df.columns):
                            result_df.iloc[row, col] = value
                            
            except Exception as e:
                self.logger.warning(f"Failed to handle merged cell region: {e}")
        
        return result_df
    
    def handle_error_cells(
        self,
        df: pd.DataFrame,
        error_handling: str = 'replace_with_nan'
    ) -> pd.DataFrame:
        """
        Handle Excel error values in DataFrame.
        
        Args:
            df: Source DataFrame
            error_handling: How to handle errors ('replace_with_nan', 'keep_as_text', 'remove_rows')
            
        Returns:
            DataFrame with error cells handled
            
        Requirements: 6.5
        """
        result_df = df.copy()
        
        # Excel error patterns
        error_patterns = [
            '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!',
            '#SPILL!', '#CALC!', '#GETTING_DATA'
        ]
        
        if error_handling == 'replace_with_nan':
            # Replace error values with NaN
            for pattern in error_patterns:
                result_df = result_df.replace(pattern, np.nan)
                
        elif error_handling == 'keep_as_text':
            # Keep as text (no action needed)
            pass
            
        elif error_handling == 'remove_rows':
            # Remove rows containing any error values
            for pattern in error_patterns:
                mask = ~result_df.astype(str).apply(lambda x: x.str.contains(pattern, na=False)).any(axis=1)
                result_df = result_df[mask]
        
        self.logger.info(f"Handled error cells with strategy: {error_handling}")
        return result_df
    
    def normalize_inconsistent_columns(
        self,
        df: pd.DataFrame,
        strategy: str = 'pad_with_nan'
    ) -> pd.DataFrame:
        """
        Handle inconsistent column counts across rows.
        
        Args:
            df: Source DataFrame
            strategy: How to handle inconsistencies ('pad_with_nan', 'truncate_to_min')
            
        Returns:
            Normalized DataFrame
            
        Requirements: 6.3
        """
        # This is mainly handled during CSV parsing, but we can clean up any issues
        result_df = df.copy()
        
        if strategy == 'pad_with_nan':
            # Ensure all rows have same number of columns (pandas handles this automatically)
            pass
        elif strategy == 'truncate_to_min':
            # Find minimum number of non-null columns per row
            min_cols = result_df.notna().sum(axis=1).min()
            if min_cols > 0 and min_cols < len(result_df.columns):
                result_df = result_df.iloc[:, :min_cols]
        
        return result_df
    
    def extract_rich_formatting_semantics(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Extract semantic meaning from rich formatting (colors, fonts, borders).
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet to analyze
            
        Returns:
            Dictionary with formatting semantics
            
        Requirements: 15.6
        """
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl not available"}
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            
            formatting_info = {
                "highlighted_cells": [],
                "bold_cells": [],
                "colored_backgrounds": {},
                "bordered_regions": [],
                "font_colors": {}
            }
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        
                        # Check for highlighting (background colors)
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != '00000000':
                            color = cell.fill.start_color.rgb
                            if color not in formatting_info["colored_backgrounds"]:
                                formatting_info["colored_backgrounds"][color] = []
                            formatting_info["colored_backgrounds"][color].append({
                                "cell": cell_ref,
                                "value": cell.value
                            })
                        
                        # Check for bold text
                        if cell.font and cell.font.bold:
                            formatting_info["bold_cells"].append({
                                "cell": cell_ref,
                                "value": cell.value
                            })
                        
                        # Check for font colors
                        if cell.font and cell.font.color and cell.font.color.rgb != '00000000':
                            color = cell.font.color.rgb
                            if color not in formatting_info["font_colors"]:
                                formatting_info["font_colors"][color] = []
                            formatting_info["font_colors"][color].append({
                                "cell": cell_ref,
                                "value": cell.value
                            })
            
            return formatting_info
            
        except Exception as e:
            self.logger.error(f"Rich formatting extraction failed: {e}")
            return {"error": str(e)}
    
    # ========================================================================
    # PRIVATE METHODS
    # ========================================================================
    
    def _extract_merged_cells(self, worksheet) -> List[MergedCellInfo]:
        """Extract merged cell information from worksheet"""
        merged_cells = []
        
        for merged_range in worksheet.merged_cells.ranges:
            # Get the top-left cell value
            top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            value = top_left_cell.value
            
            merged_cell = MergedCellInfo(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
                value=value
            )
            merged_cells.append(merged_cell)
        
        return merged_cells
    
    def _extract_formulas(self, worksheet) -> List[FormulaInfo]:
        """Extract formula information from worksheet"""
        formulas = []
        
        # Load workbook with calculated values
        workbook_with_values = openpyxl.load_workbook(worksheet.parent.path, data_only=True)
        worksheet_with_values = workbook_with_values[worksheet.title]
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    # Get calculated value from data_only workbook
                    calculated_value = worksheet_with_values.cell(cell.row, cell.column).value
                    
                    formula_info = FormulaInfo(
                        formula=cell.value,
                        calculated_value=calculated_value,
                        cell_address=f"{cell.column_letter}{cell.row}"
                    )
                    formulas.append(formula_info)
        
        return formulas
    
    def _worksheet_to_dataframe(
        self,
        worksheet,
        merged_cells: List[MergedCellInfo],
        formulas: List[FormulaInfo],
        handle_errors: bool
    ) -> pd.DataFrame:
        """Convert worksheet to DataFrame with edge case handling"""
        
        # Get worksheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 1 and max_col == 1:
            # Empty worksheet
            return pd.DataFrame()
        
        # Extract data with formula values
        data = []
        formula_dict = {f.cell_address: f.calculated_value for f in formulas}
        
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            row_data = []
            for col_idx, cell in enumerate(row, 1):
                cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                
                # Use calculated value if it's a formula
                if cell_address in formula_dict:
                    value = formula_dict[cell_address]
                else:
                    value = cell.value
                
                # Handle error values
                if handle_errors and isinstance(value, str) and value.startswith('#'):
                    value = np.nan
                
                row_data.append(value)
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Handle merged cells by replicating values
        for merged_cell in merged_cells:
            try:
                # Convert to 0-based indexing
                start_row = merged_cell.min_row - 1
                end_row = merged_cell.max_row
                start_col = merged_cell.min_col - 1
                end_col = merged_cell.max_col
                
                # Ensure indices are within bounds
                start_row = max(0, min(start_row, len(df) - 1))
                end_row = max(start_row + 1, min(end_row, len(df)))
                start_col = max(0, min(start_col, len(df.columns) - 1))
                end_col = max(start_col + 1, min(end_col, len(df.columns)))
                
                # Replicate value across merged region
                for row in range(start_row, end_row):
                    for col in range(start_col, end_col):
                        df.iloc[row, col] = merged_cell.value
                        
            except Exception as e:
                self.logger.warning(f"Failed to handle merged cell {merged_cell}: {e}")
        
        return df
    
    def _normalize_column_counts(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize inconsistent column counts"""
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Remove completely empty rows
        df = df.dropna(axis=0, how='all')
        
        return df.reset_index(drop=True)


# ============================================================================
# GLOBAL INSTANCE
# ============================================================================

# Create a global edge case handler instance
edge_case_handler = EdgeCaseHandler()